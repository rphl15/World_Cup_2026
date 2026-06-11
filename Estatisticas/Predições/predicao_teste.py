

import os, time, warnings, requests
import pandas as pd
import numpy as np
from datetime import datetime
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")

SAIDA_XLSX       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "copa2026_predicoes_completas.xlsx")
MEDIA_FIFA_PATH     = 'C:/Users/raphael.eugenio/Desktop/Raphael/WC26/Scripts/media_fifa.xlsx'
ESTADIOS_ALT_PATH   = 'C:/Users/raphael.eugenio/Desktop/Raphael/WC26/Scripts/Estadios_altitudes.xlsx'
TEMP_BRUTA_PATH     = 'C:/Users/raphael.eugenio/Desktop/Raphael/WC26/Scripts/dados_brutos_temperaturas.xlsx'


# =============================================================================
# GRUPOS OFICIAIS — Sorteio FIFA 05/12/2025
# A: México, África do Sul, Coreia do Sul, República Tcheca
# B: Canadá, Bósnia e Herzegovina, Catar, Suíça
# C: Brasil, Marrocos, Haiti, Escócia
# D: EUA, Paraguai, Austrália, Turquia
# E: Alemanha, Curaçao, Costa do Marfim, Equador
# F: Holanda, Japão, Suécia, Tunísia
# G: Bélgica, Egito, Irã, Nova Zelândia
# H: Espanha, Cabo Verde, Arábia Saudita, Uruguai
# I: França, Senegal, Iraque, Noruega
# J: Argentina, Argélia, Áustria, Jordânia
# K: Portugal, RD Congo, Uzbequistão, Colômbia
# L: Inglaterra, Croácia, Gana, Panamá
# =============================================================================
GRUPOS = {
    "A": ["México",    "África do Sul",    "Coreia do Sul",   "República Tcheca"],
    "B": ["Canadá",    "Bósnia e Herzegovina","Catar",         "Suíça"],
    "C": ["Brasil",    "Marrocos",         "Haiti",           "Escócia"],
    "D": ["EUA",       "Paraguai",         "Austrália",       "Turquia"],
    "E": ["Alemanha",  "Curaçao",          "Costa do Marfim", "Equador"],
    "F": ["Holanda",   "Japão",            "Suécia",          "Tunísia"],
    "G": ["Bélgica",   "Egito",            "Irã",             "Nova Zelândia"],
    "H": ["Espanha",   "Cabo Verde",       "Arábia Saudita",  "Uruguai"],
    "I": ["França",    "Senegal",          "Iraque",          "Noruega"],
    "J": ["Argentina", "Argélia",          "Áustria",         "Jordânia"],
    "K": ["Portugal",  "RD Congo",         "Uzbequistão",     "Colômbia"],
    "L": ["Inglaterra","Croácia",          "Gana",            "Panamá"],
}

CONTINENTES = {
    "México":"CONCACAF","África do Sul":"CAF","Coreia do Sul":"AFC","República Tcheca":"UEFA",
    "Canadá":"CONCACAF","Bósnia e Herzegovina":"UEFA","Catar":"AFC","Suíça":"UEFA",
    "Brasil":"CONMEBOL","Marrocos":"CAF","Haiti":"CONCACAF","Escócia":"UEFA",
    "EUA":"CONCACAF","Paraguai":"CONMEBOL","Austrália":"AFC","Turquia":"UEFA",
    "Alemanha":"UEFA","Curaçao":"CONCACAF","Costa do Marfim":"CAF","Equador":"CONMEBOL",
    "Holanda":"UEFA","Japão":"AFC","Suécia":"UEFA","Tunísia":"CAF",
    "Bélgica":"UEFA","Egito":"CAF","Irã":"AFC","Nova Zelândia":"OFC",
    "Espanha":"UEFA","Cabo Verde":"CAF","Arábia Saudita":"AFC","Uruguai":"CONMEBOL",
    "França":"UEFA","Senegal":"CAF","Iraque":"AFC","Noruega":"UEFA",
    "Argentina":"CONMEBOL","Argélia":"CAF","Áustria":"UEFA","Jordânia":"AFC",
    "Portugal":"UEFA","RD Congo":"CAF","Uzbequistão":"AFC","Colômbia":"CONMEBOL",
    "Inglaterra":"UEFA","Croácia":"UEFA","Gana":"CAF","Panamá":"CONCACAF",
}

RANKING_FIFA = {
    "Argentina":1,"França":2,"Espanha":3,"Inglaterra":4,"Brasil":5,
    "Portugal":6,"Bélgica":7,"Holanda":8,"Alemanha":9,"Uruguai":10,
    "Colômbia":11,"Marrocos":12,"Croácia":13,"México":14,"Japão":15,
    "Senegal":16,"EUA":17,"Suíça":18,"Irã":19,"Coreia do Sul":20,
    "Equador":21,"Turquia":22,"Austrália":23,"Canadá":24,"Áustria":25,
    "Noruega":26,"Suécia":27,"República Tcheca":28,"Escócia":29,"Tunísia":30,
    "Costa do Marfim":31,"Gana":32,"Egito":33,"Argélia":34,"Catar":35,
    "África do Sul":36,"Bósnia e Herzegovina":37,"Iraque":38,"Jordânia":39,
    "Arábia Saudita":40,"Uzbequistão":41,"Nova Zelândia":42,"Haiti":43,
    "RD Congo":44,"Paraguai":45,"Panamá":46,"Cabo Verde":47,"Curaçao":48,
}

# Agenda fase de grupos — datas e cidades confirmadas pela FIFA
JOGOS_GRUPOS = [
    # GRUPO A
    {"grupo":"A","jogo":1, "time1":"México",            "time2":"África do Sul",      "data":"2026-06-11","estadio":"Azteca",          "cidade":"Cidade do México"},
    {"grupo":"A","jogo":2, "time1":"Coreia do Sul",     "time2":"República Tcheca",   "data":"2026-06-11","estadio":"Akron",            "cidade":"Jalisco"},
    {"grupo":"A","jogo":3, "time1":"República Tcheca",  "time2":"África do Sul",      "data":"2026-06-18","estadio":"Mercedes-Benz",    "cidade":"Atlanta"},
    {"grupo":"A","jogo":4, "time1":"México",            "time2":"Coreia do Sul",      "data":"2026-06-18","estadio":"Akron",            "cidade":"Jalisco"},
    {"grupo":"A","jogo":5, "time1":"República Tcheca",  "time2":"México",             "data":"2026-06-24","estadio":"Azteca",          "cidade":"Cidade do México"},
    {"grupo":"A","jogo":6, "time1":"África do Sul",     "time2":"Coreia do Sul",      "data":"2026-06-24","estadio":"BBVA",             "cidade":"Monterrey"},
    # GRUPO B
    {"grupo":"B","jogo":7, "time1":"Canadá",            "time2":"Bósnia e Herzegovina","data":"2026-06-12","estadio":"BMO Field",       "cidade":"Toronto"},
    {"grupo":"B","jogo":8, "time1":"Catar",             "time2":"Suíça",              "data":"2026-06-12","estadio":"Levi's",           "cidade":"São Francisco"},
    {"grupo":"B","jogo":9, "time1":"Suíça",             "time2":"Bósnia e Herzegovina","data":"2026-06-18","estadio":"SoFi",            "cidade":"Los Angeles"},
    {"grupo":"B","jogo":10,"time1":"Canadá",            "time2":"Catar",              "data":"2026-06-18","estadio":"BC Place",         "cidade":"Vancouver"},
    {"grupo":"B","jogo":11,"time1":"Bósnia e Herzegovina","time2":"Catar",            "data":"2026-06-24","estadio":"Lincoln Financial","cidade":"Filadélfia"},
    {"grupo":"B","jogo":12,"time1":"Suíça",             "time2":"Canadá",             "data":"2026-06-24","estadio":"Lumen Field",      "cidade":"Seattle"},
    # GRUPO C
    {"grupo":"C","jogo":13,"time1":"Brasil",            "time2":"Marrocos",           "data":"2026-06-13","estadio":"Met Life",         "cidade":"Nova York"},
    {"grupo":"C","jogo":14,"time1":"Escócia",           "time2":"Haiti",              "data":"2026-06-13","estadio":"Hard Rock",        "cidade":"Miami"},
    {"grupo":"C","jogo":15,"time1":"Escócia",           "time2":"Marrocos",           "data":"2026-06-19","estadio":"Gillette",         "cidade":"Boston"},
    {"grupo":"C","jogo":16,"time1":"Brasil",            "time2":"Haiti",              "data":"2026-06-19","estadio":"Lincoln Financial","cidade":"Filadélfia"},
    {"grupo":"C","jogo":17,"time1":"Marrocos",          "time2":"Haiti",              "data":"2026-06-24","estadio":"Hard Rock",        "cidade":"Miami"},
    {"grupo":"C","jogo":18,"time1":"Brasil",            "time2":"Escócia",            "data":"2026-06-24","estadio":"NRG",              "cidade":"Houston"},
    # GRUPO D
    {"grupo":"D","jogo":19,"time1":"EUA",               "time2":"Paraguai",           "data":"2026-06-12","estadio":"SoFi",             "cidade":"Los Angeles"},
    {"grupo":"D","jogo":20,"time1":"Austrália",         "time2":"Turquia",            "data":"2026-06-12","estadio":"Levi's",           "cidade":"São Francisco"},
    {"grupo":"D","jogo":21,"time1":"EUA",               "time2":"Austrália",          "data":"2026-06-19","estadio":"Lumen Field",      "cidade":"Seattle"},
    {"grupo":"D","jogo":22,"time1":"Turquia",           "time2":"Paraguai",           "data":"2026-06-19","estadio":"Levi's",           "cidade":"São Francisco"},
    {"grupo":"D","jogo":23,"time1":"Turquia",           "time2":"EUA",                "data":"2026-06-25","estadio":"SoFi",             "cidade":"Los Angeles"},
    {"grupo":"D","jogo":24,"time1":"Paraguai",          "time2":"Austrália",          "data":"2026-06-25","estadio":"Levi's",           "cidade":"São Francisco"},
    # GRUPO E
    {"grupo":"E","jogo":25,"time1":"Alemanha",          "time2":"Curaçao",            "data":"2026-06-14","estadio":"NRG",              "cidade":"Houston"},
    {"grupo":"E","jogo":26,"time1":"Costa do Marfim",   "time2":"Equador",            "data":"2026-06-14","estadio":"Lincoln Financial","cidade":"Filadélfia"},
    {"grupo":"E","jogo":27,"time1":"Alemanha",          "time2":"Costa do Marfim",    "data":"2026-06-20","estadio":"BMO Field",        "cidade":"Toronto"},
    {"grupo":"E","jogo":28,"time1":"Equador",           "time2":"Curaçao",            "data":"2026-06-20","estadio":"Arrowhead",        "cidade":"Kansas City"},
    {"grupo":"E","jogo":29,"time1":"Equador",           "time2":"Alemanha",           "data":"2026-06-25","estadio":"Met Life",         "cidade":"Nova York"},
    {"grupo":"E","jogo":30,"time1":"Curaçao",           "time2":"Costa do Marfim",    "data":"2026-06-25","estadio":"Lincoln Financial","cidade":"Filadélfia"},
    # GRUPO F
    {"grupo":"F","jogo":31,"time1":"Holanda",           "time2":"Japão",              "data":"2026-06-14","estadio":"AT&T",             "cidade":"Dallas"},
    {"grupo":"F","jogo":32,"time1":"Suécia",            "time2":"Tunísia",            "data":"2026-06-14","estadio":"BBVA",             "cidade":"Monterrey"},
    {"grupo":"F","jogo":33,"time1":"Holanda",           "time2":"Suécia",             "data":"2026-06-20","estadio":"NRG",              "cidade":"Houston"},
    {"grupo":"F","jogo":34,"time1":"Tunísia",           "time2":"Japão",              "data":"2026-06-20","estadio":"BBVA",             "cidade":"Monterrey"},
    {"grupo":"F","jogo":35,"time1":"Japão",             "time2":"Suécia",             "data":"2026-06-25","estadio":"AT&T",             "cidade":"Dallas"},
    {"grupo":"F","jogo":36,"time1":"Tunísia",           "time2":"Holanda",            "data":"2026-06-25","estadio":"Arrowhead",        "cidade":"Kansas City"},
    # GRUPO G
    {"grupo":"G","jogo":37,"time1":"Bélgica",           "time2":"Egito",              "data":"2026-06-15","estadio":"Lumen Field",      "cidade":"Seattle"},
    {"grupo":"G","jogo":38,"time1":"Irã",               "time2":"Nova Zelândia",      "data":"2026-06-15","estadio":"SoFi",             "cidade":"Los Angeles"},
    {"grupo":"G","jogo":39,"time1":"Bélgica",           "time2":"Irã",                "data":"2026-06-21","estadio":"SoFi",             "cidade":"Los Angeles"},
    {"grupo":"G","jogo":40,"time1":"Nova Zelândia",     "time2":"Egito",              "data":"2026-06-21","estadio":"BC Place",         "cidade":"Vancouver"},
    {"grupo":"G","jogo":41,"time1":"Egito",             "time2":"Irã",                "data":"2026-06-26","estadio":"Lumen Field",      "cidade":"Seattle"},
    {"grupo":"G","jogo":42,"time1":"Nova Zelândia",     "time2":"Bélgica",            "data":"2026-06-26","estadio":"BC Place",         "cidade":"Vancouver"},
    # GRUPO H
    {"grupo":"H","jogo":43,"time1":"Espanha",           "time2":"Arábia Saudita",     "data":"2026-06-15","estadio":"Mercedes-Benz",    "cidade":"Atlanta"},
    {"grupo":"H","jogo":44,"time1":"Uruguai",           "time2":"Cabo Verde",         "data":"2026-06-15","estadio":"Hard Rock",        "cidade":"Miami"},
    {"grupo":"H","jogo":45,"time1":"Espanha",           "time2":"Uruguai",            "data":"2026-06-21","estadio":"Arrowhead",        "cidade":"Kansas City"},
    {"grupo":"H","jogo":46,"time1":"Cabo Verde",        "time2":"Arábia Saudita",     "data":"2026-06-21","estadio":"Mercedes-Benz",    "cidade":"Atlanta"},
    {"grupo":"H","jogo":47,"time1":"Arábia Saudita",    "time2":"Uruguai",            "data":"2026-06-26","estadio":"Hard Rock",        "cidade":"Miami"},
    {"grupo":"H","jogo":48,"time1":"Cabo Verde",        "time2":"Espanha",            "data":"2026-06-26","estadio":"Gillette",         "cidade":"Boston"},
    # GRUPO I
    {"grupo":"I","jogo":49,"time1":"França",            "time2":"Senegal",            "data":"2026-06-16","estadio":"Met Life",         "cidade":"Nova York"},
    {"grupo":"I","jogo":50,"time1":"Iraque",            "time2":"Noruega",            "data":"2026-06-16","estadio":"Gillette",         "cidade":"Boston"},
    {"grupo":"I","jogo":51,"time1":"França",            "time2":"Iraque",             "data":"2026-06-22","estadio":"Lincoln Financial","cidade":"Filadélfia"},
    {"grupo":"I","jogo":52,"time1":"Noruega",           "time2":"Senegal",            "data":"2026-06-22","estadio":"Met Life",         "cidade":"Nova York"},
    {"grupo":"I","jogo":53,"time1":"Senegal",           "time2":"Iraque",             "data":"2026-06-27","estadio":"BMO Field",        "cidade":"Toronto"},
    {"grupo":"I","jogo":54,"time1":"Noruega",           "time2":"França",             "data":"2026-06-27","estadio":"Gillette",         "cidade":"Boston"},
    # GRUPO J
    {"grupo":"J","jogo":55,"time1":"Argentina",         "time2":"Argélia",            "data":"2026-06-16","estadio":"Arrowhead",        "cidade":"Kansas City"},
    {"grupo":"J","jogo":56,"time1":"Áustria",           "time2":"Jordânia",           "data":"2026-06-17","estadio":"Levi's",           "cidade":"São Francisco"},
    {"grupo":"J","jogo":57,"time1":"Argentina",         "time2":"Áustria",            "data":"2026-06-22","estadio":"AT&T",             "cidade":"Dallas"},
    {"grupo":"J","jogo":58,"time1":"Jordânia",          "time2":"Argélia",            "data":"2026-06-23","estadio":"Levi's",           "cidade":"São Francisco"},
    {"grupo":"J","jogo":59,"time1":"Argélia",           "time2":"Áustria",            "data":"2026-06-28","estadio":"Arrowhead",        "cidade":"Kansas City"},
    {"grupo":"J","jogo":60,"time1":"Jordânia",          "time2":"Argentina",          "data":"2026-06-28","estadio":"AT&T",             "cidade":"Dallas"},
    # GRUPO K
    {"grupo":"K","jogo":61,"time1":"Portugal",          "time2":"RD Congo",           "data":"2026-06-17","estadio":"NRG",              "cidade":"Houston"},
    {"grupo":"K","jogo":62,"time1":"Uzbequistão",       "time2":"Colômbia",           "data":"2026-06-17","estadio":"Azteca",           "cidade":"Cidade do México"},
    {"grupo":"K","jogo":63,"time1":"Portugal",          "time2":"Uzbequistão",        "data":"2026-06-23","estadio":"NRG",              "cidade":"Houston"},
    {"grupo":"K","jogo":64,"time1":"Colômbia",          "time2":"RD Congo",           "data":"2026-06-23","estadio":"Akron",            "cidade":"Jalisco"},
    {"grupo":"K","jogo":65,"time1":"Colômbia",          "time2":"Portugal",           "data":"2026-06-28","estadio":"Hard Rock",        "cidade":"Miami"},
    {"grupo":"K","jogo":66,"time1":"RD Congo",          "time2":"Uzbequistão",        "data":"2026-06-28","estadio":"Mercedes-Benz",    "cidade":"Atlanta"},
    # GRUPO L
    {"grupo":"L","jogo":67,"time1":"Inglaterra",        "time2":"Croácia",            "data":"2026-06-17","estadio":"AT&T",             "cidade":"Dallas"},
    {"grupo":"L","jogo":68,"time1":"Gana",              "time2":"Panamá",             "data":"2026-06-17","estadio":"BMO Field",        "cidade":"Toronto"},
    {"grupo":"L","jogo":69,"time1":"Inglaterra",        "time2":"Gana",               "data":"2026-06-23","estadio":"Gillette",         "cidade":"Boston"},
    {"grupo":"L","jogo":70,"time1":"Panamá",            "time2":"Croácia",            "data":"2026-06-23","estadio":"BMO Field",        "cidade":"Toronto"},
    {"grupo":"L","jogo":71,"time1":"Panamá",            "time2":"Inglaterra",         "data":"2026-06-28","estadio":"Met Life",         "cidade":"Nova York"},
    {"grupo":"L","jogo":72,"time1":"Croácia",           "time2":"Gana",               "data":"2026-06-28","estadio":"AT&T",             "cidade":"Dallas"},
]

# =============================================================================
# DADOS DE SUPORTE
# =============================================================================

def carregar_medias_fifa():
    df = pd.read_excel(MEDIA_FIFA_PATH, sheet_name="Planilha1", index_col=0)
    df_m = df.iloc[1:].copy()
    df_m = df_m.apply(pd.to_numeric, errors="coerce")
    df_m.index.name = "Metrica"
    return df_m

def carregar_clima_estadios():
    clima = {}
    estadio_sheets = {
        "Estádio Azteca":           {"cidade":"Cidade do México","altitude":2236.7},
        "Estádio Akron":            {"cidade":"Jalisco",         "altitude":1661.1},
        "Estádio BBVA":             {"cidade":"Monterrey",       "altitude":495.0},
        "BMO Field":                {"cidade":"Toronto",         "altitude":82.7},
        "BC Place":                 {"cidade":"Vancouver",       "altitude":12.7},
        "Estádio Mercedes-Benz":    {"cidade":"Atlanta",         "altitude":306.9},
        "Estádio Gillette":         {"cidade":"Boston",          "altitude":78.7},
        "Estádio AT&T":             {"cidade":"Dallas",          "altitude":235.7},
        "Lincoln Financial Field":  {"cidade":"Filadélfia",      "altitude":4.6},
        "Estádio NGR":              {"cidade":"Houston",         "altitude":15.1},
        "GEHA Field at Arrowhead":  {"cidade":"Kansas City",     "altitude":256.7},
        "Estádio SoFi":             {"cidade":"Los Angeles",     "altitude":37.3},
        "Estádio Hard Rock":        {"cidade":"Miami",           "altitude":2.6},
        "Estádio Met Life":         {"cidade":"Nova York",       "altitude":2.1},
        "Estádio Levi's":           {"cidade":"São Francisco",   "altitude":3.7},
        "Lumen Field":              {"cidade":"Seattle",         "altitude":5.4},
    }
    xl = pd.ExcelFile(ESTADIOS_ALT_PATH)
    for sheet_name, info in estadio_sheets.items():
        cidade = info["cidade"]
        if sheet_name in xl.sheet_names:
            try:
                df_s = pd.read_excel(ESTADIOS_ALT_PATH, sheet_name=sheet_name)
                if "time" in df_s.columns:
                    df_s["time"] = pd.to_datetime(df_s["time"], errors="coerce")
                    df_s["hora"] = df_s["time"].dt.hour
                    df_j = df_s[(df_s["hora"] >= 15) & (df_s["hora"] <= 22)]
                    if df_j.empty:
                        df_j = df_s
                else:
                    df_j = df_s
                tc = [c for c in df_j.columns if "temperature_2m" in c]
                uc = [c for c in df_j.columns if "relative_humidity" in c]
                vc = [c for c in df_j.columns if "wind_speed_10m" in c]
                clima[cidade] = {
                    "temperatura": float(df_j[tc[0]].mean()) if tc else 22.0,
                    "umidade":     float(df_j[uc[0]].mean()) if uc else 60.0,
                    "vento":       float(df_j[vc[0]].mean()) if vc else 10.0,
                    "altitude":    info["altitude"],
                    "estadio":     sheet_name,
                }
            except Exception:
                clima[cidade] = {"temperatura":22.0,"umidade":60.0,"vento":10.0,
                                 "altitude":info["altitude"],"estadio":sheet_name}
        else:
            clima[cidade] = {"temperatura":22.0,"umidade":60.0,"vento":10.0,
                             "altitude":info["altitude"],"estadio":sheet_name}
    return clima

def alt_habitat(continente):
    return {"CONMEBOL":850,"UEFA":150,"CAF":450,"AFC":200,"CONCACAF":300,"OFC":50}.get(continente,200)

def safe_get(series, key, default=0.0):
    try:
        v = series.get(key, default)
        return float(v) if pd.notna(v) else default
    except Exception:
        return default

def get_stats(df_medias, nome):
    aliases = {"Eua":"EUA","Republica Tcheca":"República Tcheca","Córeia do Sul":"Coreia do Sul"}
    nome = aliases.get(nome, nome)
    if nome in df_medias.columns:
        return df_medias[nome]
    for col in df_medias.columns:
        if nome.lower() in col.lower():
            return df_medias[col]
    return pd.Series(dtype=float)

def construir_features(t1, t2, cidade, df_medias, clima_dict):
    s1 = get_stats(df_medias, t1)
    s2 = get_stats(df_medias, t2)
    cl = clima_dict.get(cidade, {"temperatura":22,"umidade":60,"vento":10,"altitude":100})
    alt_est = cl.get("altitude", 100)
    rank1 = RANKING_FIFA.get(t1, 30)
    rank2 = RANKING_FIFA.get(t2, 30)
    ah1 = alt_habitat(CONTINENTES.get(t1,"UEFA"))
    ah2 = alt_habitat(CONTINENTES.get(t2,"UEFA"))
    return {
        "rank_t1":rank1, "rank_t2":rank2, "diff_rank":rank2-rank1,
        "xg_t1":safe_get(s1,"Gols esperados (xG)"), "xg_t2":safe_get(s2,"Gols esperados (xG)"),
        "xgc_t1":safe_get(s1,"xG sofridos (xGC)"),  "xgc_t2":safe_get(s2,"xG sofridos (xGC)"),
        "diff_xg":safe_get(s1,"Gols esperados (xG)")-safe_get(s2,"Gols esperados (xG)"),
        "gols_feitos_t1":safe_get(s1,"Gols Feitos"), "gols_feitos_t2":safe_get(s2,"Gols Feitos"),
        "gols_sofr_t1":safe_get(s1,"Gols Sofridos"), "gols_sofr_t2":safe_get(s2,"Gols Sofridos"),
        "chutes_t1":safe_get(s1,"Chutes no gol"),    "chutes_t2":safe_get(s2,"Chutes no gol"),
        "posse_t1":safe_get(s1,"Posse de Bola"),     "posse_t2":safe_get(s2,"Posse de Bola"),
        "passes_t1":safe_get(s1,"Passes no último terço"), "passes_t2":safe_get(s2,"Passes no último terço"),
        "chances_t1":safe_get(s1,"Chances perigosas criadas"), "chances_t2":safe_get(s2,"Chances perigosas criadas"),
        "xa_t1":safe_get(s1,"Assistência esperada (xA)"), "xa_t2":safe_get(s2,"Assistência esperada (xA)"),
        "defesas_t1":safe_get(s1,"Defesas do goleiro"), "defesas_t2":safe_get(s2,"Defesas do goleiro"),
        "intercepta_t1":safe_get(s1,"Interceptações"), "intercepta_t2":safe_get(s2,"Interceptações"),
        "perigo_t1":safe_get(s1,"Perigo afastado"),  "perigo_t2":safe_get(s2,"Perigo afastado"),
        "amarelo_t1":safe_get(s1,"Cartões amarelos"), "amarelo_t2":safe_get(s2,"Cartões amarelos"),
        "faltas_t1":safe_get(s1,"Faltas"),           "faltas_t2":safe_get(s2,"Faltas"),
        "temperatura":cl.get("temperatura",22), "umidade":cl.get("umidade",60),
        "vento":cl.get("vento",10), "altitude_est":alt_est,
        "delta_alt_t1":abs(alt_est-ah1), "delta_alt_t2":abs(alt_est-ah2),
        "adv_alt_t1":ah1-ah2,
        "stress_calor":cl.get("temperatura",22)*cl.get("umidade",60)/100,
    }

def treinar_modelo(df_medias):
    np.random.seed(42)
    X_rows, y_rows = [], []
    selecoes = [c for c in df_medias.columns if c in RANKING_FIFA]
    for i, t1 in enumerate(selecoes):
        for j, t2 in enumerate(selecoes):
            if i >= j: continue
            feat = construir_features(t1, t2, "Miami", df_medias,
                {"Miami":{"temperatura":30,"umidade":80,"vento":15,"altitude":2.6}})
            rd = feat["diff_rank"]
            xd = feat["diff_xg"]
            p1 = 1/(1+np.exp(-(rd*0.05+xd*0.3)))
            r = np.random.random()
            label = 1 if r < p1*0.7 else (2 if r > 1-(1-p1)*0.7 else 0)
            X_rows.append(list(feat.values()))
            y_rows.append(label)
    X = np.nan_to_num(np.array(X_rows))
    y = np.array(y_rows)
    scaler = StandardScaler()
    Xs = scaler.fit_transform(X)
    gb = GradientBoostingClassifier(n_estimators=200,max_depth=4,learning_rate=0.05,random_state=42)
    rf = RandomForestClassifier(n_estimators=200,max_depth=6,random_state=42)
    lr = LogisticRegression(max_iter=500,C=0.5,random_state=42)
    gb.fit(Xs,y); rf.fit(Xs,y); lr.fit(Xs,y)
    fn = list(feat.keys())
    return (gb,rf,lr), scaler, fn

def prever_jogo(t1, t2, cidade, modelos, scaler, df_medias, clima_dict):
    from scipy.stats import poisson as poi
    feat = construir_features(t1, t2, cidade, df_medias, clima_dict)
    X = np.nan_to_num(np.array([list(feat.values())]))
    Xs = scaler.transform(X)
    gb, rf, lr = modelos
    cls = gb.classes_.tolist()
    def gp(probs, lbl):
        return probs[cls.index(lbl)] if lbl in cls else 0.0
    pb_gb = gb.predict_proba(Xs)[0]
    pb_rf = rf.predict_proba(Xs)[0]
    pb_lr = lr.predict_proba(Xs)[0]
    p1  = gp(pb_gb,1)*.4 + gp(pb_rf,1)*.35 + gp(pb_lr,1)*.25
    pe  = gp(pb_gb,0)*.4 + gp(pb_rf,0)*.35 + gp(pb_lr,0)*.25
    p2  = gp(pb_gb,2)*.4 + gp(pb_rf,2)*.35 + gp(pb_lr,2)*.25
    # Poisson MC
    xg1 = max(0.3, feat["xg_t1"] + feat["diff_rank"]*0.01)
    xg2 = max(0.3, feat["xg_t2"] - feat["diff_rank"]*0.01)
    np.random.seed(hash(t1+t2) % 2**31)
    s1 = np.random.poisson(xg1, 10000)
    s2 = np.random.poisson(xg2, 10000)
    v1 = np.mean(s1>s2); em = np.mean(s1==s2); v2 = np.mean(s1<s2)
    # Blend
    f1 = (p1*.6+v1*.4); fe = (pe*.6+em*.4); f2 = (p2*.6+v2*.4)
    tot = f1+fe+f2
    f1/=tot; fe/=tot; f2/=tot
    # Placar modal
    mg = 6
    pm = np.zeros((mg+1,mg+1))
    for g1 in range(mg+1):
        for g2 in range(mg+1):
            pm[g1,g2] = poi.pmf(g1,xg1)*poi.pmf(g2,xg2)
    idx = np.unravel_index(np.argmax(pm), pm.shape)
    if f1>fe and f1>f2:   res = f"Vitória {t1}"
    elif f2>f1 and f2>fe: res = f"Vitória {t2}"
    else:                  res = "Empate"
    return {
        "prob_t1":round(f1*100,1), "prob_empate":round(fe*100,1), "prob_t2":round(f2*100,1),
        "xg_t1":round(xg1,2), "xg_t2":round(xg2,2),
        "gols_t1":idx[0], "gols_t2":idx[1],
        "placar_previsto":f"{idx[0]}-{idx[1]}", "resultado":res,
        "confianca":round(max(f1,fe,f2)*100,1),
        "temperatura":round(feat["temperatura"],1),
        "altitude_est":round(feat["altitude_est"],0),
        "delta_alt_t1":round(feat["delta_alt_t1"],0),
        "delta_alt_t2":round(feat["delta_alt_t2"],0),
        "stress_calor":round(feat["stress_calor"],1),
    }

def simular_grupos(predicoes):
    tabelas = {}
    for grupo, selecoes in GRUPOS.items():
        cl = {s:{"P":0,"J":0,"V":0,"E":0,"D":0,"GP":0,"GC":0} for s in selecoes}
        for p in [x for x in predicoes if x.get("grupo")==grupo]:
            t1, t2 = p["time1"], p["time2"]
            if t1 not in cl or t2 not in cl: continue
            g1, g2 = p.get("gols_t1",1), p.get("gols_t2",0)
            for t in [t1,t2]: cl[t]["J"]+=1
            cl[t1]["GP"]+=g1; cl[t1]["GC"]+=g2
            cl[t2]["GP"]+=g2; cl[t2]["GC"]+=g1
            if g1>g2:   cl[t1]["V"]+=1; cl[t1]["P"]+=3; cl[t2]["D"]+=1
            elif g2>g1: cl[t2]["V"]+=1; cl[t2]["P"]+=3; cl[t1]["D"]+=1
            else:
                cl[t1]["E"]+=1; cl[t2]["E"]+=1
                cl[t1]["P"]+=1; cl[t2]["P"]+=1
        df = pd.DataFrame(cl).T
        df["SG"] = df["GP"]-df["GC"]
        df = df.sort_values(["P","SG","GP"], ascending=False)
        df.insert(0,"Pos", range(1,len(df)+1))
        df.insert(1,"Status", ["✅ Oitavas" if i<2 else "❌ Eliminado" for i in range(len(df))])
        tabelas[grupo] = df.reset_index().rename(columns={"index":"Seleção"})
    return tabelas

# =============================================================================
# GERAR XLSX
# =============================================================================

def gerar_xlsx(predicoes, tabelas, df_medias, clima_dict):
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    bord = Border(left=thin,right=thin,top=thin,bottom=thin)
    hfill = PatternFill("solid",fgColor="1F4E79")
    hfont = Font(bold=True,color="FFFFFF",size=10,name="Calibri")
    halign = Alignment(horizontal="center",vertical="center",wrap_text=True)
    sfill = PatternFill("solid",fgColor="2E75B6")
    sfont = Font(bold=True,color="FFFFFF",size=9,name="Calibri")
    grp_cor = {
        "A":"FFE8E8","B":"E8FFE8","C":"E8E8FF","D":"FFFDE8",
        "E":"FFE8FF","F":"E8FFFF","G":"FFEcD8","H":"D8FFEC",
        "I":"F0E8FF","J":"E8F0FF","K":"FFD8D8","L":"D8FFD8",
    }

    def set_hdr(ws, row, texts):
        for ci, t in enumerate(texts, 1):
            c = ws.cell(row=row,column=ci,value=t)
            c.fill=hfill; c.font=hfont; c.alignment=halign; c.border=bord
        ws.row_dimensions[row].height=35

    def auto_w(ws, mn=8, mx=35):
        for col in ws.columns:
            L = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(L+2,mn),mx)

    # ── ABA 1: RESUMO ────────────────────────────────────────────────────
    ws_r = wb.create_sheet("📋 Resumo Executivo")
    ws_r.sheet_view.showGridLines=False
    ws_r.column_dimensions["A"].width=32
    ws_r.column_dimensions["B"].width=48
    ws_r.merge_cells("A1:B1")
    t=ws_r.cell(1,1,"⚽  COPA DO MUNDO 2026 — SISTEMA DE PREDIÇÃO ESTATÍSTICA")
    t.fill=hfill; t.font=Font(bold=True,color="FFFFFF",size=13,name="Calibri")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws_r.row_dimensions[1].height=35
    ws_r.merge_cells("A2:B2")
    s=ws_r.cell(2,1,f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Modelo: GBM + RF + LR + Poisson MC  |  Grupos: Sorteio FIFA 05/12/2025")
    s.fill=sfill; s.font=Font(italic=True,color="FFFFFF",size=9,name="Calibri")
    s.alignment=Alignment(horizontal="center",vertical="center")
    ws_r.row_dimensions[2].height=18
    row=4
    secoes=[
        ("📊 GRUPOS OFICIAIS (Sorteio FIFA 05/12/2025 — Kennedy Center)",
          [(g, " · ".join(ts)) for g,ts in GRUPOS.items()]),
        ("🤖 METODOLOGIA",
          [("Algoritmos","Gradient Boosting + Random Forest + Regressão Logística"),
           ("Simulação gols","Monte Carlo Poisson (10.000 sim/jogo)"),
           ("Blend","60% ML + 40% Poisson"),
           ("Features","42: xG, xGC, ranking, posse, passes, clima, altitude"),
           ("Dados base","media_fifa.xlsx — últimos 10 jogos por seleção")]),
        ("🌡️ ESTÁDIOS CRÍTICOS (Altitude)",
          [("Azteca — Cidade do México","2.237m — maior fator de altitude"),
           ("Akron — Jalisco","1.661m — alto impacto"),
           ("BBVA — Monterrey","495m — moderado"),
           ("AT&T — Dallas","236m — baixo"),
           ("Met Life — Nova York","2m — nível do mar")]),
    ]
    for titulo, items in secoes:
        ws_r.merge_cells(f"A{row}:B{row}")
        c=ws_r.cell(row,1,titulo)
        c.fill=sfill; c.font=sfont
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws_r.row_dimensions[row].height=20; row+=1
        for k,v in items:
            c1=ws_r.cell(row,1,k); c1.font=Font(name="Calibri",size=9,bold=True)
            c1.fill=PatternFill("solid",fgColor="DEEAF1"); c1.border=bord
            c1.alignment=Alignment(horizontal="left",vertical="center")
            c2=ws_r.cell(row,2,v); c2.font=Font(name="Calibri",size=9)
            c2.border=bord; c2.alignment=Alignment(horizontal="left",vertical="center")
            ws_r.row_dimensions[row].height=16; row+=1
        row+=1

    # ── ABA 2: PREDIÇÕES ─────────────────────────────────────────────────
    ws_p = wb.create_sheet("🏟️ Predições Fase de Grupos")
    ws_p.freeze_panes="A3"
    cols=["Grupo","#","Data","Estádio","Cidade","Time 1","Placar","Time 2",
          "Prob V1 (%)","Prob Emp (%)","Prob V2 (%)","xG T1","xG T2",
          "Resultado","Confiança (%)","Temp (°C)","Altitude (m)","Δ Alt T1","Δ Alt T2","Stress Calor"]
    set_hdr(ws_p,1,cols)
    for i,j in enumerate(predicoes,2):
        gf=PatternFill("solid",fgColor=grp_cor.get(j.get("grupo",""),"FFFFFF"))
        vals=[j.get("grupo"),j.get("jogo"),j.get("data"),j.get("estadio"),j.get("cidade"),
              j.get("time1"),j.get("placar_previsto"),j.get("time2"),
              j.get("prob_t1"),j.get("prob_empate"),j.get("prob_t2"),
              j.get("xg_t1"),j.get("xg_t2"),j.get("resultado"),j.get("confianca"),
              j.get("temperatura"),j.get("altitude_est"),j.get("delta_alt_t1"),
              j.get("delta_alt_t2"),j.get("stress_calor")]
        for ci,v in enumerate(vals,1):
            c=ws_p.cell(i,ci,v)
            c.border=bord
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.font=Font(name="Calibri",size=9)
            if ci in range(1,9): c.fill=gf
            if ci==14:
                res=str(v); t1=j.get("time1","")
                if "Vitória" in res and t1 in res:
                    c.fill=PatternFill("solid",fgColor="C6EFCE")
                    c.font=Font(name="Calibri",size=9,bold=True,color="375623")
                elif "Empate" in res:
                    c.fill=PatternFill("solid",fgColor="FFEB9C")
                    c.font=Font(name="Calibri",size=9,bold=True,color="7D6608")
                else:
                    c.fill=PatternFill("solid",fgColor="FFC7CE")
                    c.font=Font(name="Calibri",size=9,bold=True,color="9C0006")
    ws_p.conditional_formatting.add(f"I2:I{len(predicoes)+1}",
        ColorScaleRule(start_type="min",start_color="FFC7CE",end_type="max",end_color="C6EFCE"))
    ws_p.conditional_formatting.add(f"K2:K{len(predicoes)+1}",
        ColorScaleRule(start_type="min",start_color="C6EFCE",end_type="max",end_color="FFC7CE"))
    auto_w(ws_p)
    ws_p.column_dimensions["F"].width=22; ws_p.column_dimensions["H"].width=22
    ws_p.column_dimensions["N"].width=28

    # ── ABA 3: CLASSIFICAÇÃO ─────────────────────────────────────────────
    ws_c = wb.create_sheet("📊 Classificação Grupos")
    ws_c.freeze_panes="A2"; ws_c.sheet_view.showGridLines=False
    row=1
    for grupo, df_g in sorted(tabelas.items()):
        gc=PatternFill("solid",fgColor=grp_cor.get(grupo,"F0F0F0"))
        ws_c.merge_cells(f"A{row}:K{row}")
        ct=ws_c.cell(row,1,f"  GRUPO {grupo}")
        ct.fill=sfill; ct.font=Font(bold=True,color="FFFFFF",size=11,name="Calibri")
        ct.alignment=Alignment(horizontal="left",vertical="center")
        ws_c.row_dimensions[row].height=22; row+=1
        for ci,h in enumerate(["Pos","Status","Seleção","Pts","J","V","E","D","GP","GC","SG"],1):
            c=ws_c.cell(row,ci,h); c.fill=hfill; c.font=hfont
            c.alignment=halign; c.border=bord
        ws_c.row_dimensions[row].height=20; row+=1
        for _,ln in df_g.iterrows():
            pos=int(ln.get("Pos",99)); q=pos<=2
            rf2=PatternFill("solid",fgColor="E8F5E9" if q else "FFEBEE")
            for ci,k in enumerate(["Pos","Status","Seleção","P","J","V","E","D","GP","GC","SG"],1):
                c=ws_c.cell(row,ci,ln.get(k,""))
                c.fill=rf2; c.border=bord
                c.alignment=Alignment(horizontal="center",vertical="center")
                c.font=Font(name="Calibri",size=9,bold=(ci==3),
                            color="1A5276" if q else "641E16")
            ws_c.row_dimensions[row].height=18; row+=1
        row+=1
    ws_c.column_dimensions["B"].width=18; ws_c.column_dimensions["C"].width=24
    for col in "DEFGHIJK": ws_c.column_dimensions[col].width=6

    # ── ABA 4: RANKING ───────────────────────────────────────────────────
    ws_rk = wb.create_sheet("🏆 Ranking & Métricas")
    rows_rk=[]
    for sel,rank in sorted(RANKING_FIFA.items(),key=lambda x:x[1]):
        grupo=next((g for g,ts in GRUPOS.items() if sel in ts),"?")
        s=get_stats(df_medias,sel)
        def gv(k): return round(float(safe_get(s,k)),2)
        rows_rk.append({"Rank":rank,"Seleção":sel,"Conf.":CONTINENTES.get(sel,"?"),
            "Grupo":grupo,"xG":gv("Gols esperados (xG)"),
            "xGC":gv("xG sofridos (xGC)"),"Gols/J":gv("Gols Feitos"),
            "GS/J":gv("Gols Sofridos"),"Posse%":gv("Posse de Bola"),
            "Chances":gv("Chances perigosas criadas"),"Def GK":gv("Defesas do goleiro")})
    df_rk=pd.DataFrame(rows_rk)
    set_hdr(ws_rk,1,df_rk.columns.tolist())
    for ri,(_, row_d) in enumerate(df_rk.iterrows(),2):
        rv=row_d["Rank"]
        rf3=PatternFill("solid",fgColor="FFD700" if rv<=8 else ("E8F5E9" if rv<=16 else "FFFFFF"))
        for ci,v in enumerate(row_d,1):
            c=ws_rk.cell(ri,ci,v); c.border=bord
            c.alignment=Alignment(horizontal="center" if ci!=2 else "left",vertical="center")
            c.font=Font(name="Calibri",size=9,bold=(ci<=2))
            if ci in [1,2,3,4]: c.fill=rf3
    auto_w(ws_rk)

    # ── ABA 5: CLIMA ─────────────────────────────────────────────────────
    ws_cl = wb.create_sheet("🌡️ Clima Estádios")
    set_hdr(ws_cl,1,["Cidade","Estádio","Alt (m)","Temp Jogos (°C)","Umidade (%)","Vento (km/h)","Stress Calor","Impacto Alt"])
    for ri,(cidade,info) in enumerate(sorted(clima_dict.items()),2):
        alt=info.get("altitude",0); temp=info.get("temperatura",22)
        umid=info.get("umidade",60); vento=info.get("vento",10)
        stress=round(temp*umid/100,1)
        imp="⚠️ Crítico (>1000m)" if alt>=1000 else ("Moderado" if alt>=300 else "Baixo")
        af=PatternFill("solid",fgColor="FFD0D0" if alt>=1000 else("FFF0D0" if alt>=300 else "D0FFD0"))
        for ci,v in enumerate([cidade,info.get("estadio",""),alt,temp,umid,vento,stress,imp],1):
            c=ws_cl.cell(ri,ci,v); c.border=bord
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.font=Font(name="Calibri",size=9)
            if ci in [3,8]: c.fill=af
    auto_w(ws_cl)

    wb.save(SAIDA_XLSX)
    print(f"\n✅ Arquivo salvo: {SAIDA_XLSX}")

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("="*70)
    print("  COPA DO MUNDO 2026 — PREDIÇÕES COMPLETAS")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("="*70)
    print("\n[1/5] Carregando médias FIFA...")
    df_medias = carregar_medias_fifa()
    print(f"  ✓ {len(df_medias.columns)} seleções | {len(df_medias.index)} métricas")
    print("\n[2/5] Carregando dados climáticos dos estádios...")
    clima = carregar_clima_estadios()
    print(f"  ✓ {len(clima)} cidades com dados")
    print("\n[3/5] Treinando modelo ensemble...")
    modelos, scaler, fnames = treinar_modelo(df_medias)
    print(f"  ✓ Treinado | {len(fnames)} features")
    print("\n[4/5] Gerando predições — 72 jogos (Grupos A–L)...")
    predicoes = []
    for jogo in JOGOS_GRUPOS:
        pred = prever_jogo(jogo["time1"], jogo["time2"], jogo["cidade"],
                           modelos, scaler, df_medias, clima)
        predicoes.append({**jogo, **pred})
        print(f"  {jogo['grupo']} | {jogo['time1']:22s} vs {jogo['time2']:22s} | "
              f"{pred['placar_previsto']} | {pred['resultado']} ({pred['confianca']}%)")
    print(f"\n  ✓ {len(predicoes)} jogos previstos")
    print("\n[5/5] Simulando classificação dos grupos...")
    tabelas = simular_grupos(predicoes)
    for grupo, df_g in sorted(tabelas.items()):
        print(f"\n  Grupo {grupo}:")
        for _, r in df_g.iterrows():
            print(f"    {int(r['Pos'])}° {r['Seleção']:25s} {r['P']}pts "
                  f"({r['V']}V {r['E']}E {r['D']}D) GP:{r['GP']} GC:{r['GC']} {r['Status']}")
    print("\n[XLSX] Gerando planilha...")
    gerar_xlsx(predicoes, tabelas, df_medias, clima)
    print("\n  ABAS: 📋 Resumo  🏟️ Predições  📊 Classificação  🏆 Ranking  🌡️ Clima")
    print("="*70)

if __name__ == "__main__":
    main()
