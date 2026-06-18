"""
============================================================
 COLETA DIÁRIA — Copa do Mundo 2026
============================================================
Grava na planilha Copa_2026_Estatisticas.xlsx:

  • Snapshot diário (football-data.org):
      Partidas, Classificacao, Artilheiros
  • Estatística de jogo (API-Football):
      posse de bola, finalizações, passes, xG... (1 linha por time)

São DUAS fontes porque a estatística detalhada (posse, chutes)
não existe no plano grátis da football-data.org.

--- CONFIGURAÇÃO (uma vez) ---------------------------------
  pip install requests openpyxl

  Chave 1 (snapshot):  https://www.football-data.org/client/register
  Chave 2 (stats):     https://dashboard.api-football.com/register
  Cole as duas abaixo. Se deixar uma vazia, aquela parte é pulada.

--- RODAR --------------------------------------------------
  python coleta_copa.py            -> coleta hoje
  python coleta_copa.py --force    -> refaz o snapshot do dia
============================================================
"""

import os
import sys
import datetime
import requests
from openpyxl import load_workbook

# ----------- COLE SUAS CHAVES AQUI -----------
FD_KEY = "55aebecc1b27424ca7ed3108e3b4940b"   # football-data.org
AF_KEY = "db64b60d38bafc21446c8c9c97a261fc"   # api-sports.io
PLANILHA = "Copa_2026_Estatisticas.xlsx"
# ---------------------------------------------

os.chdir(r"C:\Users\Raphael\OneDrive\Documentos")

HOJE = datetime.date.today().isoformat()

FD_BASE = "https://api.football-data.org/v4"
FD_COMP = "WC"
AF_BASE = "https://v3.football.api-sports.io"
AF_LEAGUE = 1       # FIFA World Cup
AF_SEASON = 2026
FINALIZADOS = {"FT", "AET", "PEN"}   # status de jogo encerrado


# ===================== SNAPSHOT (football-data.org) =====================
def fd_get(path, params=None):
    r = requests.get(f"{FD_BASE}/{path}",
                     headers={"X-Auth-Token": FD_KEY},
                     params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def ja_coletou_hoje(ws):
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == HOJE:
            return True
    return False


def snapshot(wb):
    # Partidas
    wsP = wb["Partidas"]
    p = 0
    for m in fd_get(f"competitions/{FD_COMP}/matches").get("matches", []):
        ft = m.get("score", {}).get("fullTime", {})
        wsP.append([
            HOJE, (m.get("utcDate") or "")[:16].replace("T", " "),
            m.get("stage", ""), (m.get("group") or "").replace("GROUP_", "Grupo "),
            m.get("homeTeam", {}).get("name", ""), ft.get("home"), ft.get("away"),
            m.get("awayTeam", {}).get("name", ""), m.get("status", ""),
            m.get("venue") or "",
        ])
        p += 1

    # Classificacao
    wsC = wb["Classificacao"]
    c = 0
    for bloco in fd_get(f"competitions/{FD_COMP}/standings").get("standings", []):
        if bloco.get("type") != "TOTAL":
            continue
        grupo = bloco.get("group") or bloco.get("stage", "")
        for t in bloco.get("table", []):
            gp, gc = t.get("goalsFor", 0), t.get("goalsAgainst", 0)
            wsC.append([
                HOJE, grupo, t.get("position"), t.get("team", {}).get("name", ""),
                t.get("playedGames"), t.get("won"), t.get("draw"), t.get("lost"),
                gp, gc, (gp or 0) - (gc or 0), t.get("points"),
            ])
            c += 1

    # Artilheiros
    wsA = wb["Artilheiros"]
    a = 0
    for i, s in enumerate(
            fd_get(f"competitions/{FD_COMP}/scorers",
                   params={"limit": 30}).get("scorers", []), start=1):
        wsA.append([
            HOJE, i, s.get("player", {}).get("name", ""),
            s.get("team", {}).get("name", ""), s.get("goals"),
            s.get("assists"), s.get("playedMatches"),
        ])
        a += 1
    return p, c, a


# ===================== ESTATÍSTICA DE JOGO (API-Football) =====================
def af_get(path, params=None):
    r = requests.get(f"{AF_BASE}/{path}",
                     headers={"x-apisports-key": AF_KEY},
                     params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def num(v):
    """Converte '56%' -> 56, '123' -> 123, None -> None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace("%", "").strip()
    try:
        return int(v)
    except (ValueError, TypeError):
        try:
            return float(v)
        except (ValueError, TypeError):
            return v


def mapa_stats(lista):
    """Lista [{type, value}] -> dict {type: value}."""
    return {s.get("type"): s.get("value") for s in (lista or [])}


def fixtures_ja_salvos(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0] is not None:
            ids.add(row[0])
    return ids


def stats_jogo(wb):
    ws = wb["EstatisticasJogo"]
    salvos = fixtures_ja_salvos(ws)

    fixtures = af_get("fixtures",
                      params={"league": AF_LEAGUE, "season": AF_SEASON}
                      ).get("response", [])
    novos = 0
    for fx in fixtures:
        fid = fx.get("fixture", {}).get("id")
        status = fx.get("fixture", {}).get("status", {}).get("short")
        if fid in salvos or status not in FINALIZADOS:
            continue

        det = af_get("fixtures/statistics", params={"fixture": fid}).get("response", [])
        if len(det) < 2:
            continue  # estatística ainda não publicada

        data_jogo = (fx.get("fixture", {}).get("date") or "")[:16].replace("T", " ")
        fase = fx.get("league", {}).get("round", "")
        home = fx.get("teams", {}).get("home", {})
        away = fx.get("teams", {}).get("away", {})
        gh = fx.get("goals", {}).get("home")
        ga = fx.get("goals", {}).get("away")
        nome = {t.get("team", {}).get("id"): t for t in det}

        def resultado(meus, outros, venceu):
            if venceu is True:
                return "V"
            if venceu is False and meus is not None and outros is not None and meus < outros:
                return "D"
            return "E"

        for time_obj, mando, gp, gc in (
                (home, "Casa", gh, ga), (away, "Fora", ga, gh)):
            tid = time_obj.get("id")
            bloco = nome.get(tid, {})
            st = mapa_stats(bloco.get("statistics"))
            adversario = (away if mando == "Casa" else home).get("name", "")
            ws.append([
                HOJE, fid, data_jogo, fase, time_obj.get("name", ""), adversario,
                mando, gp, gc, resultado(gp, gc, time_obj.get("winner")),
                num(st.get("Ball Possession")),
                num(st.get("Total Shots")),
                num(st.get("Shots on Goal")),
                num(st.get("Shots off Goal")),
                num(st.get("Blocked Shots")),
                num(st.get("Shots insidebox")),
                num(st.get("Shots outsidebox")),
                num(st.get("Corner Kicks")),
                num(st.get("Offsides")),
                num(st.get("Fouls")),
                num(st.get("Goalkeeper Saves")),
                num(st.get("Total passes")),
                num(st.get("Passes accurate")),
                num(st.get("Passes %")),
                num(st.get("Yellow Cards")),
                num(st.get("Red Cards")),
                num(st.get("expected_goals")),
            ])
        novos += 1
    return novos


# ===================== MAIN =====================
def main():
    forcar = "--force" in sys.argv
    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        print(f"Nao achei '{PLANILHA}'. Deixe na mesma pasta do script.")
        sys.exit(1)

    # 1) Snapshot diario
    if FD_KEY and FD_KEY != "COLE_CHAVE_FOOTBALL_DATA":
        if forcar or not ja_coletou_hoje(wb["Partidas"]):
            try:
                p, c, a = snapshot(wb)
                print(f"Snapshot {HOJE}: {p} partidas, {c} linhas de tabela, {a} artilheiros.")
            except requests.HTTPError as e:
                print(f"Snapshot falhou: {e} (confira a chave football-data.org).")
        else:
            print(f"Snapshot de hoje ja existe. Use --force para refazer.")
    else:
        print("Sem chave football-data.org -> snapshot pulado.")

    # 2) Estatistica de jogo
    if AF_KEY and AF_KEY != "COLE_CHAVE_API_FOOTBALL":
        try:
            n = stats_jogo(wb)
            print(f"Estatistica de jogo: {n} partida(s) nova(s) adicionada(s).")
        except requests.HTTPError as e:
            print(f"Estatistica falhou: {e} (confira a chave/season da API-Football).")
    else:
        print("Sem chave API-Football -> estatistica de jogo pulada.")

    wb.save(PLANILHA)
    print("Planilha salva.")


if __name__ == "__main__":
    main()
