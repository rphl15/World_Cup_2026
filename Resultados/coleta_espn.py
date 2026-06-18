"""
============================================================
 COLETA ESPN — estatísticas de jogo da Copa 2026
============================================================
Fonte: API pública da ESPN (sem chave, sem Cloudflare).
Grava na aba 'Estatisticas' TODAS as estatísticas que a ESPN
mostra por jogo (posse, finalizações, chutes no gol, escanteios,
faltas, cartões...), em formato LONGO: 1 linha por estatística,
para cada time. Depois é só filtrar no Excel o que quiser.

  pip install requests openpyxl
  python coleta_espn.py
============================================================
"""
import os
import time
import datetime
import requests
from openpyxl import load_workbook

PLANILHA = "Copa_2026_Estatisticas.xlsx"

try:
    os.chdir(r"C:\Users\Raphael\OneDrive\Documentos")
except OSError:
    pass

LIGA = "fifa.world"   # Copa do Mundo FIFA (masculino) na ESPN
ROOT = f"https://site.api.espn.com/apis/site/v2/sports/soccer/{LIGA}"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
INICIO = datetime.date(2026, 6, 11)   # 1º dia da Copa
HOJE = datetime.date.today()
HOJE_STR = HOJE.isoformat()
PAUSA = 0.8


def get(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def ja_salvos(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0] is not None:
            ids.add(str(row[0]))
    return ids


def coletar():
    wb = load_workbook(PLANILHA)
    if "Estatisticas" not in wb.sheetnames:
        ws = wb.create_sheet("Estatisticas")
        ws.append(["data_coleta", "event_id", "data_jogo", "fase", "mando",
                   "time", "adversario", "gols_pro", "gols_contra",
                   "estatistica", "valor"])
    ws = wb["Estatisticas"]
    salvos = ja_salvos(ws)

    # 1) varre os dias da Copa e junta os jogos encerrados
    jogos = {}
    dia = INICIO
    total_eventos = 0
    while dia <= HOJE:
        try:
            data = get(f"{ROOT}/scoreboard", params={"dates": dia.strftime("%Y%m%d")})
        except requests.HTTPError:
            dia += datetime.timedelta(days=1)
            continue
        for ev in data.get("events", []):
            total_eventos += 1
            comp = (ev.get("competitions") or [{}])[0]
            if not comp.get("status", {}).get("type", {}).get("completed"):
                continue
            cs = comp.get("competitors", [])
            info = {}
            for c in cs:
                info[c.get("homeAway")] = {
                    "id": c.get("team", {}).get("id"),
                    "nome": c.get("team", {}).get("displayName", ""),
                    "gols": c.get("score"),
                }
            nota = ""
            if comp.get("notes"):
                nota = comp["notes"][0].get("headline", "")
            jogos[str(ev.get("id"))] = {
                "data": (ev.get("date") or "")[:16].replace("T", " "),
                "fase": nota,
                "home": info.get("home", {}),
                "away": info.get("away", {}),
            }
        dia += datetime.timedelta(days=1)
        time.sleep(PAUSA)

    print(f"Eventos encontrados: {total_eventos} | jogos encerrados: {len(jogos)}")
    if total_eventos == 0:
        print("Nenhum evento veio da ESPN. O nome da liga pode ter mudado.")
        print("Me avise que eu ajusto (testo outros nomes de liga).")
        wb.save(PLANILHA)
        return

    # 2) para cada jogo novo, pega TODAS as estatísticas
    novos = 0
    for eid, j in jogos.items():
        if eid in salvos:
            continue
        time.sleep(PAUSA)
        try:
            summ = get(f"{ROOT}/summary", params={"event": eid})
        except requests.HTTPError:
            continue
        times = (summ.get("boxscore") or {}).get("teams", [])
        if not times:
            continue

        home, away = j["home"], j["away"]
        por_id = {}
        for t in times:
            tid = t.get("team", {}).get("id")
            por_id[tid] = t.get("statistics", [])

        for lado, mando, adv in ((home, "Casa", away), (away, "Fora", home)):
            stats = por_id.get(lado.get("id"), [])
            for st in stats:
                nome = st.get("label") or st.get("name", "")
                valor = st.get("displayValue", st.get("value"))
                ws.append([HOJE_STR, eid, j["data"], j["fase"], mando,
                           lado.get("nome", ""), adv.get("nome", ""),
                           home.get("gols") if mando == "Casa" else away.get("gols"),
                           away.get("gols") if mando == "Casa" else home.get("gols"),
                           nome, valor])
        novos += 1
        print(f"  + {home.get('nome')} {home.get('gols')}-{away.get('gols')} {away.get('nome')}")

    wb.save(PLANILHA)
    print(f"OK: {novos} jogo(s) novo(s) na aba 'Estatisticas'. Planilha salva.")


if __name__ == "__main__":
    try:
        coletar()
    except requests.HTTPError as ex:
        print(f"Erro HTTP: {ex}")
        print("Me mande essa mensagem que eu ajusto.")
