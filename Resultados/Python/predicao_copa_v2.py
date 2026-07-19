"""
============================================================
 PREDIÇÃO DO MATA-MATA — Copa do Mundo 2026  (VERSÃO 2)
============================================================
O que mudou em relação à v1:

  [NOVO] 1. ATUALIZAÇÃO AUTOMÁTICA: antes de prever, o script
         verifica a idade dos dados (coluna 'data_coleta').
         Se estiverem velhos, ele roda o 'coleta_espn.py'
         sozinho (se estiver na mesma pasta) e recarrega.
  [NOVO] 2. CHAVEAMENTO REAL: quando a ESPN já publicou os
         confrontos do mata-mata na aba 'Partidas', o script
         usa os confrontos REAIS em vez de reconstruir a
         classificação dos grupos (que erra desempates da
         FIFA, como confronto direto e alocação de terceiros).
  [NOVO] 3. RESULTADOS REAIS DO MATA-MATA: jogos FINISHED da
         aba 'Partidas' entram como 'JOGADO'. Empate (decisão
         nos pênaltis) é resolvido olhando quem aparece na
         fase seguinte.
  [NOVO] 4. AVISO DE DADOS VELHOS: a data da coleta aparece
         no terminal e dentro da aba de predições.

Ordem de uso (agora pode ser só um comando):
   python predicao_copa_v2.py     -> coleta (se preciso) + prevê

   pip install openpyxl
============================================================
"""
import os
import sys
import math
import random
import subprocess
from collections import defaultdict
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------
# CONFIGURAÇÃO
# ----------------------------------------------------------
PLANILHA = "Copa_2026_Estatisticas.xlsx"
COLETOR = "coleta_espn (1).py"        # script que baixa os dados da ESPN
AUTO_COLETA = True                # True = roda o coletor sozinho se os dados estiverem velhos
MAX_IDADE_DIAS = 1                # dados com mais de 1 dia = velhos (Copa tem jogo todo dia)

try:
    os.chdir(r"C:\Users\Raphael\Downloads")
except OSError:
    pass

N_SIMULACOES = 20000
# A fase de grupos termina em 27/06 à noite e o mata-mata começa 28/06.
INICIO_MATA_MATA = "2026-06-28 12:00"
SEED = 42
random.seed(SEED)

NORMALIZA_NOME = {
    "Türkiye": "Turkey",
    "Cape Verde": "Cape Verde Islands",
}

PESO_GOLS_REAIS = 0.6
XG_POR_CHUTE_NO_ALVO = 0.31
ESCALA_FORCA = 320

DEFAULT_ELO = 1700
ELO_INICIAL = {
    # Favoritos
    "Argentina": 2085, "France": 2080, "Spain": 2075, "Brazil": 2045,
    "England": 2030, "Portugal": 2010, "Netherlands": 1995, "Germany": 1985,
    # Fortes
    "Belgium": 1935, "Croatia": 1900, "Uruguay": 1895, "Colombia": 1890,
    "Turkey": 1880, "Morocco": 1875, "Switzerland": 1855, "Japan": 1840,
    "Senegal": 1825, "Norway": 1820, "United States": 1820, "Mexico": 1805,
    # Intermediários
    "Ecuador": 1790, "South Korea": 1790, "Austria": 1790, "Iran": 1785,
    "Sweden": 1780, "Algeria": 1775, "Egypt": 1770, "Czechia": 1770,
    "Scotland": 1765, "Ivory Coast": 1760, "Australia": 1755, "Paraguay": 1740,
    "Canada": 1740, "Ghana": 1725, "Bosnia-Herzegovina": 1715, "Congo DR": 1705,
    # Mais fracos
    "Qatar": 1700, "Tunisia": 1690, "Panama": 1685, "Saudi Arabia": 1670,
    "Cape Verde Islands": 1655, "South Africa": 1650, "Uzbekistan": 1650,
    "Iraq": 1640, "Jordan": 1620, "New Zealand": 1600, "Curaçao": 1575,
    "Haiti": 1560,
}

# ==========================================================
# CHAVEAMENTO OFICIAL (usado só como RESERVA, quando a ESPN
# ainda não publicou os confrontos reais na aba 'Partidas')
# ==========================================================
R32 = [
    (73, ("R", "A"), ("R", "B")),
    (74, ("W", "E"), ("T", "D")),
    (75, ("W", "F"), ("R", "C")),
    (76, ("W", "C"), ("R", "F")),
    (77, ("W", "I"), ("T", "F")),
    (78, ("R", "E"), ("R", "I")),
    (79, ("W", "A"), ("T", "E")),
    (80, ("W", "L"), ("T", "K")),
    (81, ("W", "D"), ("T", "B")),
    (82, ("W", "G"), ("T", "I")),
    (83, ("W", "H"), ("R", "J")),
    (84, ("W", "B"), ("T", "J")),
    (85, ("R", "K"), ("R", "L")),
    (86, ("W", "J"), ("R", "H")),
    (87, ("W", "K"), ("T", "L")),
    (88, ("R", "D"), ("R", "G")),
]
R16 = [(89, 74, 77), (90, 73, 75), (91, 76, 78), (92, 79, 80),
       (93, 83, 84), (94, 81, 82), (95, 86, 88), (96, 85, 87)]
QF = [(97, 89, 90), (98, 91, 92), (99, 93, 94), (100, 95, 96)]
SF = [(101, 97, 98), (102, 99, 100)]
FINAL_NUM = 104
TERCEIRO_NUM = 103

# fase (ESPN) -> nome bonito, e ordem das fases
FASES_KM = ["LAST_32", "LAST_16", "QUARTER_FINALS",
            "SEMI_FINALS", "THIRD_PLACE", "FINAL"]
NOME_FASE_ESPN = {"LAST_32": "Oitavas (R32)", "LAST_16": "16-avos (R16)",
                  "QUARTER_FINALS": "Quartas", "SEMI_FINALS": "Semifinal",
                  "THIRD_PLACE": "3º lugar", "FINAL": "FINAL"}
# número do 1º jogo de cada fase (para mapear linhas reais -> números)
PRIMEIRO_NUM = {"LAST_32": 73, "LAST_16": 89, "QUARTER_FINALS": 97,
                "SEMI_FINALS": 101, "THIRD_PLACE": 103, "FINAL": 104}


def fase_do_jogo(num):
    if 73 <= num <= 88:
        return "Oitavas (R32)"
    if 89 <= num <= 96:
        return "16-avos (R16)"
    if 97 <= num <= 100:
        return "Quartas"
    if num in (101, 102):
        return "Semifinal"
    if num == 103:
        return "3º lugar"
    return "FINAL"


# ==========================================================
# 0) [NOVO] FRESCOR DOS DADOS + COLETA AUTOMÁTICA
# ==========================================================
def data_da_coleta(wb):
    """Maior 'data_coleta' encontrada na aba Partidas."""
    ws = wb["Partidas"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {n: i for i, n in enumerate(linhas[0])}
    datas = [str(r[h["data_coleta"]])[:10] for r in linhas[1:]
             if r[h["data_coleta"]]]
    return max(datas) if datas else None


def garantir_dados_atualizados():
    """Confere a idade dos dados. Se velhos, roda o coleta_espn.py."""
    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        wb = None

    coleta = data_da_coleta(wb) if wb else None
    hoje = date.today()
    idade = None
    if coleta:
        idade = (hoje - datetime.strptime(coleta, "%Y-%m-%d").date()).days

    precisa_coletar = (wb is None) or (idade is None) or (idade > MAX_IDADE_DIAS)

    if precisa_coletar:
        print("=" * 60)
        if idade is not None:
            print(f"!! DADOS VELHOS: última coleta em {coleta} ({idade} dias atrás).")
        else:
            print("!! Planilha não encontrada ou sem data de coleta.")
        if AUTO_COLETA and os.path.exists(COLETOR):
            print(f">> Rodando '{COLETOR}' para atualizar os dados...")
            print("=" * 60)
            ret = subprocess.run([sys.executable, COLETOR])
            if ret.returncode != 0:
                print(f"!! '{COLETOR}' terminou com erro. Predição vai usar os dados antigos.")
        else:
            print(f"!! Rode 'python {COLETOR}' primeiro, senão a predição")
            print("!! vai mostrar jogos ERRADOS (times já eliminados etc).")
        print("=" * 60)

    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        print(f"Não achei '{PLANILHA}'. Deixe na mesma pasta do script.")
        return None, None
    return wb, data_da_coleta(wb)


# ==========================================================
# 1) LEITURA DOS DADOS
# ==========================================================
def norm(nome):
    if nome is None:
        return None
    nome = str(nome).strip()
    if not nome:
        return None
    return NORMALIZA_NOME.get(nome, nome)


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("%", "").strip())
    except ValueError:
        return None


def ler_painel(wb):
    ws = wb["Painel"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    jogos = []
    for r in linhas[1:]:
        if r[h["gols_pro"]] is None or r[h["gols_contra"]] is None:
            continue
        jogos.append({
            "data": str(r[h["data_jogo"]]),
            "time": norm(r[h["time"]]),
            "adv": norm(r[h["adversario"]]),
            "gp": int(r[h["gols_pro"]]),
            "gc": int(r[h["gols_contra"]]),
            "chutes_alvo": _num(r[h.get("fin_no_gol")]) if "fin_no_gol" in h else None,
        })
    return jogos


def ler_grupos(wb):
    ws = wb["Classificacao"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    grupo = {}
    for r in linhas[1:]:
        sel = norm(r[h["selecao"]])
        g = str(r[h["grupo"]]).replace("Group ", "").replace("Grupo ", "").strip()
        if sel:
            grupo[sel] = g
    return grupo


def ler_partidas_mata_mata(wb):
    """[NOVO] Lê os jogos do mata-mata da aba 'Partidas' com os times REAIS
    que a ESPN publicou. Devolve: fase -> lista de jogos em ordem de data."""
    ws = wb["Partidas"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    por_fase = defaultdict(list)
    for r in linhas[1:]:
        fase = str(r[h["fase"]]) if r[h["fase"]] else ""
        if fase not in FASES_KM:
            continue
        por_fase[fase].append({
            "data": str(r[h["data_jogo"]]),
            "a": norm(r[h["mandante"]]),
            "b": norm(r[h["visitante"]]),
            "ga": r[h["gols_mand"]],
            "gb": r[h["gols_vis"]],
            "status": str(r[h["status"]]) if r[h["status"]] else "",
        })
    for fase in por_fase:
        por_fase[fase].sort(key=lambda j: j["data"])
    return por_fase


def ler_datas_mata_mata(wb):
    ws = wb["Partidas"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    datas = defaultdict(list)
    for r in linhas[1:]:
        fase = NOME_FASE_ESPN.get(str(r[h["fase"]]))
        if fase:
            datas[fase].append(str(r[h["data_jogo"]]))
    for k in datas:
        datas[k].sort()
    return datas


# ==========================================================
# 2) CLASSIFICAÇÃO DOS GRUPOS (RESERVA — só quando a ESPN
#    ainda não publicou os confrontos reais do mata-mata)
# ==========================================================
def classificar(jogos, grupo):
    tab = defaultdict(lambda: dict(J=0, V=0, E=0, D=0, GP=0, GC=0, Pts=0))
    for j in jogos:
        if j["data"][:16] >= INICIO_MATA_MATA:
            continue
        s = tab[j["time"]]
        s["J"] += 1; s["GP"] += j["gp"]; s["GC"] += j["gc"]
        if j["gp"] > j["gc"]:
            s["V"] += 1; s["Pts"] += 3
        elif j["gp"] < j["gc"]:
            s["D"] += 1
        else:
            s["E"] += 1; s["Pts"] += 1

    def chave(t):
        s = tab[t]
        return (s["Pts"], s["GP"] - s["GC"], s["GP"])

    posicoes, terceiros = {}, []
    for g in sorted(set(grupo.values())):
        times = sorted([t for t in tab if grupo.get(t) == g],
                       key=chave, reverse=True)
        if len(times) >= 1:
            posicoes[("W", g)] = times[0]
        if len(times) >= 2:
            posicoes[("R", g)] = times[1]
        if len(times) >= 3:
            terceiros.append((times[2], g, tab[times[2]]))
    return tab, posicoes, terceiros


def melhores_terceiros(terceiros):
    ordenados = sorted(
        terceiros,
        key=lambda x: (x[2]["Pts"], x[2]["GP"] - x[2]["GC"], x[2]["GP"]),
        reverse=True,
    )
    melhores = ordenados[:8]
    return {g: t for (t, g, _) in melhores}, [t for (t, g, _) in melhores]


def montar_r32_reserva(jogos, grupo):
    """Monta o R32 pela classificação calculada (aproximação).
    AVISO: não conhece todos os desempates da FIFA nem a tabela
    oficial de alocação de terceiros — use só enquanto a ESPN
    não publica os confrontos reais."""
    _, posicoes, terceiros = classificar(jogos, grupo)
    melhores, ranking_terceiros = melhores_terceiros(terceiros)

    usados = set()

    def resolve(vaga):
        tipo, g = vaga
        if tipo in ("W", "R"):
            return posicoes.get((tipo, g))
        t = melhores.get(g)
        if t:
            usados.add(t)
        return t

    slots = []
    pendentes = []   # vagas de terceiro sem dono direto
    for (num, va, vb) in R32:
        a, b = resolve(va), resolve(vb)
        slots.append([num, a, b, va, vb])
        if a is None and va[0] == "T":
            pendentes.append((len(slots) - 1, 1))
        if b is None and vb[0] == "T":
            pendentes.append((len(slots) - 1, 2))

    # [NOVO] terceiro do grupo pedido não classificou -> preenche a vaga
    # com o melhor terceiro ainda não usado (aproximação da tabela FIFA)
    sobras = [t for t in ranking_terceiros if t not in usados]
    for (idx, pos) in pendentes:
        if sobras:
            slots[idx][pos] = sobras.pop(0)

    return [(num, a, b) for (num, a, b, _, _) in slots]


def montar_r32(wb, jogos, grupo, partidas_km):
    """[NOVO] Prefere os confrontos REAIS da ESPN; senão, usa a reserva."""
    reais = partidas_km.get("LAST_32", [])
    completos = [j for j in reais if j["a"] and j["b"]]
    if len(completos) == len(R32) and len(R32) == len(reais):
        print(">> Chaveamento R32: usando confrontos REAIS da ESPN (aba Partidas).")
        return [(73 + i, j["a"], j["b"]) for i, j in enumerate(reais)], True
    print(">> Chaveamento R32: ESPN ainda não publicou os confrontos.")
    print("   Usando classificação calculada (APROXIMAÇÃO — desempates da")
    print("   FIFA e alocação de terceiros podem diferir).")
    return montar_r32_reserva(jogos, grupo), False


def parear_fases_reais(partidas_km):
    """[NOVO] Para cada fase depois do R32, se a ESPN já publicou TODOS os
    confrontos com os dois times, devolve os pares reais (em ordem de data).
    fase -> lista [(num, a, b), ...]"""
    pares = {}
    tamanho = {"LAST_16": 8, "QUARTER_FINALS": 4, "SEMI_FINALS": 2,
               "THIRD_PLACE": 1, "FINAL": 1}
    for fase, n in tamanho.items():
        reais = partidas_km.get(fase, [])
        completos = [j for j in reais if j["a"] and j["b"]]
        if len(reais) == n and len(completos) == n:
            num0 = PRIMEIRO_NUM[fase]
            pares[fase] = [(num0 + i, j["a"], j["b"]) for i, j in enumerate(reais)]
    return pares


# ==========================================================
# 3) FORÇA DAS SELEÇÕES (Elo + ataque/defesa Poisson com xG)
# ==========================================================
def calcular_forcas(jogos):
    times = set()
    for j in jogos:
        times.add(j["time"]); times.add(j["adv"])
    elo = {t: float(ELO_INICIAL.get(t, DEFAULT_ELO)) for t in times}
    K = 20
    vistos = set()
    for j in sorted(jogos, key=lambda x: x["data"]):
        par = tuple(sorted([j["time"], j["adv"]]) + [j["data"]])
        if par in vistos:
            continue
        vistos.add(par)
        a, b = j["time"], j["adv"]
        ea = 1 / (1 + 10 ** ((elo[b] - elo[a]) / 400))
        sa = 1.0 if j["gp"] > j["gc"] else (0.0 if j["gp"] < j["gc"] else 0.5)
        mult = 1 + math.log(1 + abs(j["gp"] - j["gc"]))
        delta = K * mult * (sa - ea)
        elo[a] += delta
        elo[b] -= delta

    def gol_ajustado(g, chutes_alvo):
        if chutes_alvo is None:
            return float(g)
        xg = chutes_alvo * XG_POR_CHUTE_NO_ALVO
        return PESO_GOLS_REAIS * g + (1 - PESO_GOLS_REAIS) * xg

    media_liga = (sum(gol_ajustado(j["gp"], j["chutes_alvo"]) for j in jogos)
                  / max(len(jogos), 1))

    elo_medio = sum(elo.values()) / len(elo)
    prior = {}
    for t in elo:
        s = math.exp((elo[t] - elo_medio) / ESCALA_FORCA)
        prior[t] = {"atk": math.sqrt(s), "dfs": 1 / math.sqrt(s)}

    atk = {t: prior[t]["atk"] for t in elo}
    dfs = {t: prior[t]["dfs"] for t in elo}
    PESO_PRIOR = 3.0

    for _ in range(60):
        novo_atk, novo_dfs = {}, {}
        for t in elo:
            jg = [j for j in jogos if j["time"] == t]
            gp = sum(gol_ajustado(j["gp"], j["chutes_alvo"]) for j in jg)
            esperado_atk = sum(media_liga * dfs[j["adv"]] for j in jg)
            gc = sum(gol_ajustado(j["gc"], None) for j in jg)
            esperado_dfs = sum(media_liga * atk[j["adv"]] for j in jg)
            novo_atk[t] = ((gp + PESO_PRIOR * prior[t]["atk"] * media_liga) /
                           (esperado_atk + PESO_PRIOR * media_liga))
            novo_dfs[t] = ((gc + PESO_PRIOR * prior[t]["dfs"] * media_liga) /
                           (esperado_dfs + PESO_PRIOR * media_liga))
        atk, dfs = novo_atk, novo_dfs

    return {"elo": dict(elo), "atk": atk, "dfs": dfs,
            "media_liga": media_liga, "prior": prior}


# ==========================================================
# 4) PREVISÃO DE UM JOGO (Poisson)
# ==========================================================
def _poisson(lmbda, k):
    return math.exp(-lmbda) * lmbda ** k / math.factorial(k)


def prever_jogo(a, b, F, max_gols=8):
    la = F["media_liga"] * F["atk"].get(a, 1) * F["dfs"].get(b, 1)
    lb = F["media_liga"] * F["atk"].get(b, 1) * F["dfs"].get(a, 1)
    pa = pe = pb = 0.0
    melhor_p, placar = -1, (0, 0)
    for x in range(max_gols + 1):
        for y in range(max_gols + 1):
            p = _poisson(la, x) * _poisson(lb, y)
            if x > y:
                pa += p
            elif x == y:
                pe += p
            else:
                pb += p
            if p > melhor_p:
                melhor_p, placar = p, (x, y)
    forca_a = F["atk"].get(a, 1) / F["dfs"].get(a, 1)
    forca_b = F["atk"].get(b, 1) / F["dfs"].get(b, 1)
    p_pen_a = forca_a / (forca_a + forca_b)
    p_avanca_a = pa + pe * p_pen_a
    return {"la": la, "lb": lb, "pa": pa, "pe": pe, "pb": pb,
            "placar": placar, "p_avanca_a": p_avanca_a}


# ==========================================================
# 5) [MELHORADO] RESULTADO REAL de um jogo do mata-mata
# ==========================================================
def quem_avancou_pela_fase_seguinte(a, b, fase_espn, partidas_km):
    """Empate (pênaltis): descobre quem avançou olhando quem aparece
    nos confrontos reais da fase seguinte."""
    if fase_espn == "SEMI_FINALS":
        # quem está na FINAL avançou; quem está no 3º lugar perdeu
        for j in partidas_km.get("FINAL", []):
            if a in (j["a"], j["b"]):
                return a
            if b in (j["a"], j["b"]):
                return b
        for j in partidas_km.get("THIRD_PLACE", []):
            if a in (j["a"], j["b"]):
                return b
            if b in (j["a"], j["b"]):
                return a
        return None
    proxima = {"LAST_32": "LAST_16", "LAST_16": "QUARTER_FINALS",
               "QUARTER_FINALS": "SEMI_FINALS"}.get(fase_espn)
    if not proxima:
        return None
    for j in partidas_km.get(proxima, []):
        if a in (j["a"], j["b"]):
            return a
        if b in (j["a"], j["b"]):
            return b
    return None


def resultado_real(a, b, partidas_km, fase_espn=None):
    """Procura o jogo (a x b) nos dados REAIS da aba 'Partidas'.
    Devolve o vencedor, ou None se ainda não foi jogado / não dá pra saber."""
    fases = [fase_espn] if fase_espn else FASES_KM
    for fase in fases:
        for j in partidas_km.get(fase, []):
            if j["a"] and j["b"] and {j["a"], j["b"]} == {a, b}:
                if j["status"] != "FINISHED":
                    return None
                if j["ga"] is None or j["gb"] is None:
                    return None
                ga, gb = int(j["ga"]), int(j["gb"])
                if ga > gb:
                    return j["a"]
                if gb > ga:
                    return j["b"]
                # empate = decidido nos pênaltis -> olha a fase seguinte
                return quem_avancou_pela_fase_seguinte(a, b, fase, partidas_km)
    return None


# ==========================================================
# 6) MONTE CARLO
# ==========================================================
def simular_torneio(slots_r32, F, partidas_km, pares_reais, n=N_SIMULACOES):
    cache = {}

    def prob(a, b):
        key = (a, b)
        if key not in cache:
            cache[key] = prever_jogo(a, b, F)["p_avanca_a"]
        return cache[key]

    fase_num = {}
    for (num, _, _) in R16:
        fase_num[num] = "LAST_16"
    for (num, _, _) in QF:
        fase_num[num] = "QUARTER_FINALS"
    for (num, _, _) in SF:
        fase_num[num] = "SEMI_FINALS"

    # confrontos reais já publicados sobrepõem a árvore teórica
    reais_por_num = {}
    for fase, lista in pares_reais.items():
        for (num, a, b) in lista:
            reais_por_num[num] = (a, b)

    contagem = defaultdict(lambda: defaultdict(int))
    for _ in range(n):
        venc = {}
        for (num, a, b) in slots_r32:
            r = resultado_real(a, b, partidas_km, "LAST_32")
            if r is None:
                r = a if random.random() < prob(a, b) else b
            venc[num] = r
            contagem[a]["Oitavas (R32)"]
            contagem[b]["Oitavas (R32)"]
            contagem[r]["16-avos (R16)"] += 1
        for jogos_fase, nome in [(R16, "Quartas"), (QF, "Semifinal"), (SF, "FINAL")]:
            for (num, x, y) in jogos_fase:
                if num in reais_por_num:
                    a, b = reais_por_num[num]
                else:
                    a, b = venc[x], venc[y]
                r = resultado_real(a, b, partidas_km, fase_num[num])
                if r is None:
                    r = a if random.random() < prob(a, b) else b
                venc[num] = r
                if nome != "FINAL":
                    contagem[r][nome] += 1
        a, b = venc[101], venc[102]
        if FINAL_NUM in reais_por_num:
            a, b = reais_por_num[FINAL_NUM]
        contagem[a].setdefault("Finalista", 0); contagem[a]["Finalista"] += 1
        contagem[b].setdefault("Finalista", 0); contagem[b]["Finalista"] += 1
        campeao = resultado_real(a, b, partidas_km, "FINAL")
        if campeao is None:
            campeao = a if random.random() < prob(a, b) else b
        contagem[campeao].setdefault("Campeão", 0)
        contagem[campeao]["Campeão"] += 1

    probs = {}
    for t, d in contagem.items():
        probs[t] = {
            "R16": d.get("16-avos (R16)", 0) / n,
            "QF": d.get("Quartas", 0) / n,
            "SF": d.get("Semifinal", 0) / n,
            "Final": d.get("Finalista", 0) / n,
            "Campeao": d.get("Campeão", 0) / n,
        }
    return probs


# ==========================================================
# 7) CHAVEAMENTO "FAVORITO AVANÇA"
# ==========================================================
def chave_favoritos(slots_r32, F, partidas_km, pares_reais):
    venc, jogos_prev = {}, []

    fase_num = {num: "LAST_32" for (num, _, _) in R32}
    for (num, _, _) in R16:
        fase_num[num] = "LAST_16"
    for (num, _, _) in QF:
        fase_num[num] = "QUARTER_FINALS"
    for (num, _, _) in SF:
        fase_num[num] = "SEMI_FINALS"
    fase_num[TERCEIRO_NUM] = "THIRD_PLACE"
    fase_num[FINAL_NUM] = "FINAL"

    reais_por_num = {}
    for fase, lista in pares_reais.items():
        for (num, a, b) in lista:
            reais_por_num[num] = (a, b)

    def joga(num, a, b):
        pr = prever_jogo(a, b, F)
        real = resultado_real(a, b, partidas_km, fase_num.get(num))
        if real:
            r, status = real, "JOGADO"
        else:
            r = a if pr["p_avanca_a"] >= 0.5 else b
            status = "previsto"
        jogos_prev.append({"num": num, "fase": fase_do_jogo(num),
                           "a": a, "b": b, "pr": pr, "venc": r, "status": status})
        return r

    for (num, a, b) in slots_r32:
        venc[num] = joga(num, a, b)
    for grupo_fase in (R16, QF, SF):
        for (num, x, y) in grupo_fase:
            if num in reais_por_num:
                a, b = reais_por_num[num]     # confronto REAL da ESPN
            else:
                a, b = venc[x], venc[y]       # derivado da árvore
            venc[num] = joga(num, a, b)
    # 3º lugar
    if TERCEIRO_NUM in reais_por_num:
        p1, p2 = reais_por_num[TERCEIRO_NUM]
    else:
        perd_sf = []
        for (num, x, y) in SF:
            a, b = venc[x], venc[y]
            perd_sf.append(a if venc[num] == b else b)
        p1, p2 = perd_sf[0], perd_sf[1]
    venc[TERCEIRO_NUM] = joga(TERCEIRO_NUM, p1, p2)
    # final
    if FINAL_NUM in reais_por_num:
        fa, fb = reais_por_num[FINAL_NUM]
    else:
        fa, fb = venc[101], venc[102]
    venc[FINAL_NUM] = joga(FINAL_NUM, fa, fb)
    return jogos_prev, venc


# ==========================================================
# 8) ESCREVER AS ABAS
# ==========================================================
AZUL = "1F4E78"; AZUL_CLARO = "DDEBF7"; VERDE = "C6EFCE"; CINZA = "F2F2F2"
LARANJA = "FCE4D6"


def estilo_cabecalho(ws, ncols, linha=1):
    fill = PatternFill("solid", start_color=AZUL)
    fonte = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=linha, column=c)
        cell.fill = fill; cell.font = fonte
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bd
    ws.row_dimensions[linha].height = 22


def estilo_cabecalho_intervalo(ws, linha, col0, ncols):
    fill = PatternFill("solid", start_color=AZUL)
    fonte = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for c in range(col0, col0 + ncols):
        cell = ws.cell(linha, c)
        cell.fill = fill; cell.font = fonte
        cell.alignment = Alignment(horizontal="center", vertical="center")


def cor_gradiente(x):
    r = int(0xDD + (0x1F - 0xDD) * x)
    g = int(0xEB + (0x4E - 0xEB) * x)
    b = int(0xF7 + (0x78 - 0xF7) * x)
    return f"{r:02X}{g:02X}{b:02X}"


def pct(x):
    return f"{100*x:.1f}%"


def escrever_predicoes(wb, jogos_prev, probs, coleta, chave_real):
    if "Predicoes_MataMata" in wb.sheetnames:
        del wb["Predicoes_MataMata"]
    ws = wb.create_sheet("Predicoes_MataMata")

    arial = Font(name="Arial", size=11)
    bold = Font(name="Arial", size=11, bold=True)

    ws["A1"] = "CHAVEAMENTO PREVISTO — COPA 2026"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=AZUL)
    origem = ("confrontos REAIS da ESPN" if chave_real
              else "classificação CALCULADA (aproximação)")
    ws["A2"] = (f"Dados coletados em: {coleta} | Chaveamento: {origem} | "
                "'JOGADO' = resultado real | 'previsto' = predição (Poisson + Elo + xG).")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")

    # [NOVO] aviso visível se os dados estiverem velhos
    idade = (date.today() - datetime.strptime(coleta, "%Y-%m-%d").date()).days
    if idade > MAX_IDADE_DIAS:
        ws["A3"] = (f"ATENÇÃO: dados de {idade} dias atrás! Rode o coleta_espn.py "
                    "para atualizar, senão as predições ficam erradas.")
        ws["A3"].font = Font(name="Arial", size=10, bold=True, color="C00000")
        ws["A3"].fill = PatternFill("solid", start_color=LARANJA)

    cab = ["Jogo", "Fase", "Mandante", "Visitante", "Placar provável",
           "P(Mandante)", "Empate", "P(Visitante)", "Quem avança", "Status"]
    ws.append([])
    ws.append(cab)
    linha_cab = ws.max_row
    estilo_cabecalho(ws, len(cab), linha_cab)

    for jp in jogos_prev:
        pr = jp["pr"]
        ws.append([
            jp["num"], jp["fase"], jp["a"], jp["b"],
            f"{pr['placar'][0]}-{pr['placar'][1]}",
            pct(pr["pa"]), pct(pr["pe"]), pct(pr["pb"]),
            jp["venc"], jp["status"],
        ])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = arial
            ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r, 3).alignment = Alignment(horizontal="left")
        ws.cell(r, 4).alignment = Alignment(horizontal="left")
        ws.cell(r, 9).font = bold
        if jp["status"] == "JOGADO":
            for c in range(1, len(cab) + 1):
                ws.cell(r, c).fill = PatternFill("solid", start_color=CINZA)
        if jp["num"] == FINAL_NUM:
            for c in range(1, len(cab) + 1):
                ws.cell(r, c).fill = PatternFill("solid", start_color=VERDE)
                ws.cell(r, c).font = bold

    larguras = [6, 16, 20, 20, 14, 12, 9, 12, 20, 10]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{linha_cab + 1}"

    col0 = len(cab) + 2
    ws.cell(linha_cab - 1, col0, "PROBABILIDADES (Monte Carlo)").font = bold
    cab2 = ["Seleção", "Oitavas→R16", "→Quartas", "→Semi", "→Final", "Campeão"]
    for i, t in enumerate(cab2):
        ws.cell(linha_cab, col0 + i, t)
    estilo_cabecalho_intervalo(ws, linha_cab, col0, len(cab2))

    ranking = sorted(probs.items(), key=lambda kv: kv[1]["Campeao"], reverse=True)
    r = linha_cab + 1
    for time, p in ranking:
        vals = [time, pct(p["R16"]), pct(p["QF"]), pct(p["SF"]),
                pct(p["Final"]), pct(p["Campeao"])]
        for i, v in enumerate(vals):
            cell = ws.cell(r, col0 + i, v)
            cell.font = arial
            cell.alignment = Alignment(horizontal="center" if i else "left")
        intensidade = min(p["Campeao"] * 3, 1)
        if intensidade > 0.02:
            ws.cell(r, col0 + 5).fill = PatternFill(
                "solid", start_color=cor_gradiente(intensidade))
        r += 1
    larg2 = [20, 12, 11, 10, 10, 11]
    for i, w in enumerate(larg2):
        ws.column_dimensions[get_column_letter(col0 + i)].width = w


def escrever_proximos(wb, jogos_prev, datas_fase):
    if "Proximos_Jogos" in wb.sheetnames:
        del wb["Proximos_Jogos"]
    ws = wb.create_sheet("Proximos_Jogos")

    arial = Font(name="Arial", size=11)
    bold = Font(name="Arial", size=11, bold=True)
    ws["A1"] = "PRÓXIMOS JOGOS + PREDIÇÃO"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=AZUL)
    ws["A2"] = "Jogos do mata-mata que ainda não foram jogados, em ordem de fase."
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")

    cab = ["Data prevista", "Fase", "Jogo", "Placar provável",
           "P(time 1)", "Empate", "P(time 2)", "Favorito", "Confiança"]
    ws.append([]); ws.append(cab)
    linha_cab = ws.max_row
    estilo_cabecalho(ws, len(cab), linha_cab)

    contador_fase = defaultdict(int)
    proximos = [jp for jp in jogos_prev if jp["status"] != "JOGADO"]
    for jp in proximos:
        pr = jp["pr"]
        fase = jp["fase"]
        lista = datas_fase.get(fase, [])
        idx = contador_fase[fase]
        data = lista[idx][:16] if idx < len(lista) else "a definir"
        contador_fase[fase] += 1
        conf = max(pr["pa"], pr["pb"])
        ws.append([
            data, fase, f"{jp['a']} x {jp['b']}",
            f"{pr['placar'][0]}-{pr['placar'][1]}",
            pct(pr["pa"]), pct(pr["pe"]), pct(pr["pb"]),
            jp["venc"], pct(conf),
        ])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = arial
            ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r, 3).alignment = Alignment(horizontal="left")
        ws.cell(r, 8).font = bold
        if conf >= 0.6:
            ws.cell(r, 9).fill = PatternFill("solid", start_color=VERDE)

    larguras = [16, 16, 30, 14, 11, 9, 11, 20, 11]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{linha_cab + 1}"


# ==========================================================
# MAIN
# ==========================================================
def main():
    # [NOVO] 1º passo: garantir dados frescos (roda o coletor se preciso)
    wb, coleta = garantir_dados_atualizados()
    if wb is None:
        return
    print(f"Dados da planilha: coleta de {coleta}")

    jogos = ler_painel(wb)
    grupo = ler_grupos(wb)
    datas_fase = ler_datas_mata_mata(wb)
    partidas_km = ler_partidas_mata_mata(wb)
    print(f"Jogos lidos do Painel: {len(jogos)//2}")

    # [NOVO] chaveamento real da ESPN sempre que existir
    slots_r32, chave_real = montar_r32(wb, jogos, grupo, partidas_km)
    pares_reais = parear_fases_reais(partidas_km)
    if pares_reais:
        print(">> Fases com confrontos reais publicados:",
              ", ".join(NOME_FASE_ESPN[f] for f in pares_reais))

    vazios = [(num, a, b) for (num, a, b) in slots_r32 if a is None or b is None]
    for (num, a, b) in vazios:
        print(f"  ! vaga vazia no jogo {num}: {a} x {b}")

    F = calcular_forcas(jogos)
    print("\nTop 8 por Elo:")
    for t, e in sorted(F["elo"].items(), key=lambda kv: kv[1], reverse=True)[:8]:
        print(f"  {t:<22} Elo {e:6.0f}")

    jogos_prev, _ = chave_favoritos(slots_r32, F, partidas_km, pares_reais)
    probs = simular_torneio(slots_r32, F, partidas_km, pares_reais)

    escrever_predicoes(wb, jogos_prev, probs, coleta, chave_real)
    escrever_proximos(wb, jogos_prev, datas_fase)
    wb.save(PLANILHA)

    print("\n=== CAMINHO PREVISTO ATÉ A FINAL ===")
    for jp in jogos_prev:
        if jp["num"] in (101, 102, 103, 104):
            print(f"  {jp['fase']:<12} {jp['a']} x {jp['b']}  -> {jp['venc']}")
    campeao = max(probs.items(), key=lambda kv: kv[1]["Campeao"])
    print(f"\nFavorito ao título: {campeao[0]} ({pct(campeao[1]['Campeao'])})")
    print("\nAbas atualizadas: 'Predicoes_MataMata' e 'Proximos_Jogos'. Planilha salva.")


if __name__ == "__main__":
    main()
