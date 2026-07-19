# -*- coding: utf-8 -*-
"""
============================================================
 PREDICAO DA FINAL — Copa do Mundo 2026
============================================================
Rode este script DEPOIS das semifinais.

O que ele faz:
  1. (Opcional) roda o coleta_espn.py para atualizar o Painel.
  2. Le os resultados das semis direto da aba 'Painel'
     (resolve penaltis lendo a nota "X advance on penalties").
  3. Descobre sozinho os finalistas e quem joga o 3o lugar.
  4. Roda o SEU modelo (Elo + ataque/defesa Poisson com proxy
     de xG) do predicao_copa_v2.py e preve FINAL + 3o lugar.
  5. Escreve a aba 'Predicoes_Final' e mostra no terminal.

Uso:
    python predicao_final.py

    pip install openpyxl
============================================================
"""
import os
import sys
import random
import importlib.util
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------- CONFIGURACAO -----------------
PLANILHA = r"C:\Users\Raphael\Downloads\wdc\Copa_2026_Estatisticas.xlsx"
MODELO = r"C:\Users\Raphael\Downloads\wdc\predicao_copa_v2.py"      # de onde vem o seu modelo (Elo/Poisson/xG)
COLETOR = r"C:\Users\Raphael\Downloads\wdc\coleta_espn (1).py"      # coletor da ESPN
AUTO_COLETA = True                  # True = atualiza o Painel antes de prever
N_SIM = 100000                      # simulacoes Monte Carlo
SEED = 42

# Os dois confrontos de semifinal (ajuste se precisar):
SEMIS = [("France", "Spain"), ("Argentina", "England")]

# OVERRIDE MANUAL (opcional): se o auto-detector nao achar o
# resultado no Painel, preencha quem venceu cada semi aqui.
#   Ex.:  ("France", "Spain"): "Spain"
VENCEDOR_MANUAL = {
    ("France", "Spain"): None,
    ("Argentina", "England"): None,
}
# ------------------------------------------------

AZUL = "1F4E78"
random.seed(SEED)


def carregar_modelo():
    """Importa as funcoes do seu predicao_copa_v2.py (mesmo Elo/Poisson/xG)."""
    if not os.path.exists(MODELO):
        print(f"!! Nao achei '{MODELO}'. Deixe este script na mesma pasta dele.")
        sys.exit(1)
    argv = sys.argv
    sys.argv = ["x"]  # evita que flags atrapalhem o import
    spec = importlib.util.spec_from_file_location("pred", MODELO)
    pred = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(pred)
    sys.argv = argv
    return pred


def atualizar_dados():
    """Roda o coletor da ESPN, se existir, para deixar o Painel fresco."""
    if AUTO_COLETA and os.path.exists(COLETOR):
        print(f">> Atualizando dados com '{COLETOR}'...")
        import subprocess
        subprocess.run([sys.executable, COLETOR])
    elif AUTO_COLETA:
        print(f"!! '{COLETOR}' nao encontrado — usando o Painel como esta.")


def vencedor_da_semi(wb, a, b):
    """Le a aba Painel e devolve (vencedor, perdedor, placar) da semi a x b.
    Resolve empate (penaltis) pela nota 'X advance on penalties'.
    Devolve None se a partida ainda nao aparece no Painel."""
    ws = wb["Painel"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {n: i for i, n in enumerate(linhas[0])}
    for r in linhas[1:]:
        time = r[h["time"]]
        adv = r[h["adversario"]]
        if {time, adv} != {a, b}:
            continue
        gp, gc = r[h["gols_pro"]], r[h["gols_contra"]]
        if gp is None or gc is None:
            continue
        gp, gc = int(gp), int(gc)
        placar = f"{time} {gp}-{gc} {adv}"
        if gp > gc:
            return time, adv, placar
        if gc > gp:
            return adv, time, placar
        # empate -> penaltis: olha a nota na coluna 'fase'
        nota = str(r[h["fase"]]) if r[h["fase"]] else ""
        for t in (a, b):
            if t.lower() in nota.lower() and "penal" in nota.lower():
                return t, (b if t == a else a), placar + f"  ({nota})"
        # empate sem nota clara -> nao da pra saber
        return None
    return None


def resolver_finalistas(wb):
    """Devolve (finalistas, terceiro_lugar) a partir das semis."""
    vencedores, perdedores, placares = [], [], []
    for (a, b) in SEMIS:
        manual = VENCEDOR_MANUAL.get((a, b))
        if manual:
            venc = manual
            perd = b if manual == a else a
            placares.append(f"{a} x {b}: {venc} (manual)")
        else:
            res = vencedor_da_semi(wb, a, b)
            if res is None:
                print(f"!! Ainda nao achei o resultado de {a} x {b} no Painel.")
                print(f"   Rode o coleta_espn.py, ou preencha VENCEDOR_MANUAL.")
                return None, None, None
            venc, perd, placar = res
            placares.append(placar)
        vencedores.append(venc)
        perdedores.append(perd)
    return vencedores, perdedores, placares


def prob_avanca(pred, F, a, b):
    return pred.prever_jogo(a, b, F)["p_avanca_a"]


def main():
    atualizar_dados()

    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        print(f"Nao achei '{PLANILHA}'. Deixe na mesma pasta do script.")
        return

    pred = carregar_modelo()

    finalistas, terceiro, placares = resolver_finalistas(wb)
    if finalistas is None:
        return
    f1, f2 = finalistas
    t1, t2 = terceiro

    print("\nResultados das semis:")
    for p in placares:
        print("  ", p)
    print(f"\nFINAL previsto: {f1} x {f2}")
    print(f"3o lugar:       {t1} x {t2}")

    # forcas pelo Painel (SEU modelo)
    jogos = pred.ler_painel(wb)
    F = pred.calcular_forcas(jogos)

    prev_final = pred.prever_jogo(f1, f2, F)
    prev_terc = pred.prever_jogo(t1, t2, F)

    def linha_jogo(nome, a, b, pr):
        pav_a = pr["p_avanca_a"]
        fav = a if pav_a >= 0.5 else b
        print(f"\n== {nome}: {a} x {b}")
        print(f"   placar provavel: {pr['placar'][0]}-{pr['placar'][1]}")
        print(f"   90 min:  {a} {pr['pa']*100:.1f}%  |  empate {pr['pe']*100:.1f}%  |  {b} {pr['pb']*100:.1f}%")
        print(f"   vence:   {a} {pav_a*100:.1f}%  |  {b} {(1-pav_a)*100:.1f}%   -> favorito {fav}")
        return fav

    campeao_prev = linha_jogo("FINAL", f1, f2, prev_final)
    linha_jogo("3o LUGAR", t1, t2, prev_terc)
    print(f"\n>>> CAMPEAO previsto: {campeao_prev}")

    # ----- escreve aba Predicoes_Final -----
    if "Predicoes_Final" in wb.sheetnames:
        del wb["Predicoes_Final"]
    ws = wb.create_sheet("Predicoes_Final", 0)
    fill = PatternFill("solid", start_color=AZUL)
    fhead = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    bold = Font(name="Arial", size=11, bold=True)
    arial = Font(name="Arial", size=11)

    ws["A1"] = "PREDICAO - FINAL COPA 2026"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=AZUL)
    ws["A2"] = "Modelo: Elo + ataque/defesa Poisson (proxy xG) do predicao_copa_v2.py. Base: aba Painel."
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")

    cab = ["Jogo", "Confronto", "Placar provavel", "P(time1) 90min",
           "Empate 90min", "P(time2) 90min", "Vence time1", "Vence time2", "Favorito"]
    ws.append([]); ws.append(cab)
    lc = ws.max_row
    for c in range(1, len(cab) + 1):
        cell = ws.cell(lc, c); cell.fill = fill; cell.font = fhead
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for nome, a, b, pr in [("FINAL", f1, f2, prev_final),
                           ("3o lugar", t1, t2, prev_terc)]:
        pav = pr["p_avanca_a"]
        ws.append([nome, f"{a} x {b}", f"{pr['placar'][0]}-{pr['placar'][1]}",
                   f"{pr['pa']*100:.1f}%", f"{pr['pe']*100:.1f}%", f"{pr['pb']*100:.1f}%",
                   f"{pav*100:.1f}%", f"{(1-pav)*100:.1f}%",
                   a if pav >= 0.5 else b])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = arial
            ws.cell(r, c).alignment = Alignment(horizontal="center" if c > 2 else "left")
        ws.cell(r, 9).font = bold

    ws.append([])
    ws.append([f"CAMPEAO PREVISTO: {campeao_prev}"])
    ws.cell(ws.max_row, 1).font = Font(name="Arial", size=12, bold=True, color="1F7A1F")

    larguras = [10, 22, 15, 14, 13, 14, 12, 12, 14]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(PLANILHA)
    print(f"\nAba 'Predicoes_Final' criada. Planilha salva: {PLANILHA}")


if __name__ == "__main__":
    main()
