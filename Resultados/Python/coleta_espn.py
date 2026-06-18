"""
============================================================
 COLETA ESPN + PAINEL — Copa do Mundo 2026
============================================================
Toda vez que você roda, ele:
  1. Coleta os jogos NOVOS da ESPN (sem chave, sem bloqueio)
     -> grava na aba 'Estatisticas' (formato longo, guarda tudo)
  2. Reconstrói a aba 'Painel' (formato largo, 1 linha por time
     por jogo, já em número, com resultado V/E/D) -> pronto pra
     analisar e fazer predição.

  pip install requests openpyxl
  python coleta_espn.py
============================================================
"""
import os
import time
import datetime
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PLANILHA = "Copa_2026_Estatisticas.xlsx"

try:
    os.chdir(r"C:\Users\raphael.eugenio\Downloads")
except OSError:
    pass

LIGA = "fifa.world"
ROOT = f"https://site.api.espn.com/apis/site/v2/sports/soccer/{LIGA}"
HEADERS = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}
INICIO = datetime.date(2026, 6, 11)
HOJE = datetime.date.today()
HOJE_STR = HOJE.isoformat()
PAUSA = 0.8

# ---- mapeamento ESPN -> nome de coluna (aba Painel) ----
MAP = [
    ("Possession", "posse_%"), ("SHOTS", "finalizacoes"), ("ON GOAL", "fin_no_gol"),
    ("Blocked Shots", "fin_bloqueadas"), ("Corner Kicks", "escanteios"),
    ("Offsides", "impedimentos"), ("Fouls", "faltas"), ("Yellow Cards", "amarelos"),
    ("Red Cards", "vermelhos"), ("Saves", "defesas_gk"), ("Passes", "passes"),
    ("Accurate Passes", "passes_certos"), ("Pass Completion %", "passes_%"),
    ("Crosses", "cruzamentos"), ("Accurate Crosses", "cruzamentos_certos"),
    ("Tackles", "desarmes"), ("Interceptions", "interceptacoes"),
    ("Clearances", "cortes"), ("Long Balls", "bolas_longas"),
]
CTX = ["data_jogo", "fase", "mando", "time", "adversario",
       "gols_pro", "gols_contra", "resultado"]
PAINEL_HEADERS = CTX + [dest for _, dest in MAP]


def get(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def _num(v):
    if v is None:
        return None
    s = str(v).replace("%", "").strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return v


def ja_salvos(ws):
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if row[0] is not None:
            ids.add(str(row[0]))
    return ids


# ===================== 1) COLETA (formato longo) =====================
def coletar(wb):
    if "Estatisticas" not in wb.sheetnames:
        ws = wb.create_sheet("Estatisticas")
        ws.append(["data_coleta", "event_id", "data_jogo", "fase", "mando",
                   "time", "adversario", "gols_pro", "gols_contra",
                   "estatistica", "valor"])
    ws = wb["Estatisticas"]
    salvos = ja_salvos(ws)

    jogos, dia, total = {}, INICIO, 0
    while dia <= HOJE:
        try:
            data = get(f"{ROOT}/scoreboard", params={"dates": dia.strftime("%Y%m%d")})
        except requests.HTTPError:
            dia += datetime.timedelta(days=1)
            continue
        for ev in data.get("events", []):
            total += 1
            comp = (ev.get("competitions") or [{}])[0]
            if not comp.get("status", {}).get("type", {}).get("completed"):
                continue
            info = {}
            for c in comp.get("competitors", []):
                info[c.get("homeAway")] = {
                    "id": c.get("team", {}).get("id"),
                    "nome": c.get("team", {}).get("displayName", ""),
                    "gols": c.get("score"),
                }
            nota = comp["notes"][0].get("headline", "") if comp.get("notes") else ""
            jogos[str(ev.get("id"))] = {
                "data": (ev.get("date") or "")[:16].replace("T", " "),
                "fase": nota, "home": info.get("home", {}), "away": info.get("away", {}),
            }
        dia += datetime.timedelta(days=1)
        time.sleep(PAUSA)

    print(f"Eventos: {total} | jogos encerrados: {len(jogos)}")
    if total == 0:
        print("Nenhum evento veio da ESPN — me avise que ajusto o nome da liga.")
        return

    novos = 0
    for eid, j in jogos.items():
        if eid in salvos:
            continue
        time.sleep(PAUSA)
        try:
            summ = get(f"{ROOT}/summary", params={"event": eid})
        except requests.HTTPError:
            continue
        times = (summ.get("boxscore") or {}).get("teams", [])
        if not times:
            continue
        home, away = j["home"], j["away"]
        por_id = {t.get("team", {}).get("id"): t.get("statistics", []) for t in times}
        for lado, mando, adv in ((home, "Casa", away), (away, "Fora", home)):
            for st in por_id.get(lado.get("id"), []):
                nome = st.get("label") or st.get("name", "")
                ws.append([HOJE_STR, eid, j["data"], j["fase"], mando,
                           lado.get("nome", ""), adv.get("nome", ""),
                           home.get("gols") if mando == "Casa" else away.get("gols"),
                           away.get("gols") if mando == "Casa" else home.get("gols"),
                           nome, st.get("displayValue", st.get("value"))])
        novos += 1
        print(f"  + {home.get('nome')} {home.get('gols')}-{away.get('gols')} {away.get('nome')}")
    print(f"Jogos novos coletados: {novos}")


# ===================== 2) PAINEL (formato largo) =====================
def build_painel(wb):
    rows = list(wb["Estatisticas"].iter_rows(values_only=True))
    if len(rows) < 2:
        return 0
    h = {name: i for i, name in enumerate(rows[0])}
    jogos = {}
    for r in rows[1:]:
        key = (r[h["event_id"]], r[h["mando"]])
        d = jogos.setdefault(key, {
            "data_jogo": r[h["data_jogo"]], "fase": r[h["fase"]],
            "mando": r[h["mando"]], "time": r[h["time"]],
            "adversario": r[h["adversario"]], "gols_pro": r[h["gols_pro"]],
            "gols_contra": r[h["gols_contra"]], "event_id": r[h["event_id"]], "stats": {},
        })
        d["stats"][r[h["estatistica"]]] = r[h["valor"]]

    if "Painel" in wb.sheetnames:
        del wb["Painel"]
    ws = wb.create_sheet("Painel")
    ws.append(PAINEL_HEADERS)

    def resultado(gp, gc):
        try:
            gp, gc = int(gp), int(gc)
        except (TypeError, ValueError):
            return ""
        return "V" if gp > gc else ("D" if gp < gc else "E")

    ordem = sorted(jogos.values(),
                   key=lambda d: (str(d["data_jogo"]), str(d["event_id"]),
                                  0 if d["mando"] == "Casa" else 1))
    for d in ordem:
        linha = [d["data_jogo"], d["fase"], d["mando"], d["time"], d["adversario"],
                 _num(d["gols_pro"]), _num(d["gols_contra"]),
                 resultado(d["gols_pro"], d["gols_contra"])]
        for src_name, _ in MAP:
            linha.append(_num(d["stats"].get(src_name)))
        ws.append(linha)

    fill = PatternFill("solid", start_color="1F4E78")
    fhead = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(PAINEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill; cell.font = fhead
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bd
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PAINEL_HEADERS))}1"
    larg = [18, 14, 7, 18, 18, 9, 11, 10] + [11] * len(MAP)
    for i, w in enumerate(larg, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return len(ordem)


# ===================== MAIN =====================
def main():
    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        print(f"Nao achei '{PLANILHA}'. Deixe na mesma pasta do script.")
        return
    try:
        coletar(wb)
    except requests.HTTPError as ex:
        print(f"Erro HTTP na coleta: {ex}")
    n = build_painel(wb)
    wb.save(PLANILHA)
    print(f"Painel reconstruido: {n} linhas. Planilha salva.")


if __name__ == "__main__":
    main()
