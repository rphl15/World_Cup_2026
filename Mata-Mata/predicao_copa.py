"""
============================================================
 PREDIÇÃO DO MATA-MATA — Copa do Mundo 2026
============================================================
Companheiro do seu 'coleta_espn.py'. A ordem de uso é:

   1) python coleta_espn.py     -> atualiza 'Estatisticas' e 'Painel'
   2) python predicao_copa.py   -> cria/atualiza as abas de predição

O que este programa faz, toda vez que você roda:
   1. Lê as estatísticas da aba 'Painel' (todos os jogos já jogados).
   2. Calcula a força de cada seleção:
        - Elo (rating que sobe/desce a cada jogo, ajustado pelo placar)
        - ataque/defesa estilo Poisson, com proxy de xG (chutes no gol)
   3. Monta o chaveamento oficial do mata-mata (oitavas -> final),
      preenchendo os times a partir da classificação dos grupos.
   4. Simula o torneio (Monte Carlo) e calcula a chance de cada
      seleção chegar em cada fase e de ser campeã.
   5. Escreve duas abas novas:
        - 'Predicoes_MataMata' : chaveamento previsto + probabilidades
        - 'Proximos_Jogos'     : próximos jogos com a predição de cada um

   pip install openpyxl
   python predicao_copa.py
============================================================
"""
import os
import math
import random
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------
# CONFIGURAÇÃO
# ----------------------------------------------------------
PLANILHA = "Copa_2026_Estatisticas.xlsx"

try:
    os.chdir(r"C:\Users\Raphael\Downloads")
except OSError:
    pass

N_SIMULACOES = 20000          # quantas vezes o torneio é simulado
# A fase de grupos termina em 28/06 de madrugada e o mata-mata começa
# 28/06 à noite. Usamos um corte ao MEIO-DIA de 28/06 para separar os dois.
INICIO_MATA_MATA = "2026-06-28 12:00"   # jogos a partir daqui = mata-mata
SEED = 42                     # deixa o resultado reproduzível
random.seed(SEED)

# Nomes diferentes entre abas -> padroniza para o nome usado em 'Classificacao'
NORMALIZA_NOME = {
    "Türkiye": "Turkey",
    "Cape Verde": "Cape Verde Islands",
}

# Mistura entre gols reais e proxy de xG (0 = só xG, 1 = só gols reais)
PESO_GOLS_REAIS = 0.6
# Quantos "gols esperados" vale um chute no alvo (proxy simples de xG)
XG_POR_CHUTE_NO_ALVO = 0.31
# Quanto a diferença de Elo "pesa" na força. MAIOR = mais zebra/equilíbrio,
# MENOR = favoritos mais dominantes. Use isto para calibrar o realismo.
ESCALA_FORCA = 320

# ----------------------------------------------------------
# ELO INICIAL (força ANTES da Copa) — ancorado no ranking FIFA de 11/06/2026.
# É o "prior": sem isso, o modelo acha que quem ganhou de grupo fraco é favorito.
# Pode ajustar os números à vontade; o que importa é a ordem relativa.
# ----------------------------------------------------------
DEFAULT_ELO = 1700
ELO_INICIAL = {
    # Favoritos
    "Argentina": 2085, "France": 2080, "Spain": 2075, "Brazil": 2045,
    "England": 2030, "Portugal": 2010, "Netherlands": 1995, "Germany": 1985,
    # Fortes
    "Belgium": 1935, "Croatia": 1900, "Uruguay": 1895, "Colombia": 1890,
    "Turkey": 1880, "Morocco": 1875, "Switzerland": 1855, "Japan": 1840,
    "Senegal": 1825, "Norway": 1820, "United States": 1820, "Mexico": 1805,
    # Intermediários
    "Ecuador": 1790, "South Korea": 1790, "Austria": 1790, "Iran": 1785,
    "Sweden": 1780, "Algeria": 1775, "Egypt": 1770, "Czechia": 1770,
    "Scotland": 1765, "Ivory Coast": 1760, "Australia": 1755, "Paraguay": 1740,
    "Canada": 1740, "Ghana": 1725, "Bosnia-Herzegovina": 1715, "Congo DR": 1705,
    # Mais fracos
    "Qatar": 1700, "Tunisia": 1690, "Panama": 1685, "Saudi Arabia": 1670,
    "Cape Verde Islands": 1655, "South Africa": 1650, "Uzbekistan": 1650,
    "Iraq": 1640, "Jordan": 1620, "New Zealand": 1600, "Curaçao": 1575,
    "Haiti": 1560,
}

# ==========================================================
# CHAVEAMENTO OFICIAL DA COPA 2026 (48 seleções)
# ==========================================================
# Cada vaga é um código:
#   ('W', 'A') = 1º (Winner) do Grupo A
#   ('R', 'B') = 2º (Runner-up) do Grupo B
#   ('T', 'D') = 3º (Third) do Grupo D  -> um dos 8 melhores terceiros
#
# (num_jogo, vaga_casa, vaga_fora)
R32 = [
    (73, ("R", "A"), ("R", "B")),
    (74, ("W", "E"), ("T", "D")),
    (75, ("W", "F"), ("R", "C")),
    (76, ("W", "C"), ("R", "F")),
    (77, ("W", "I"), ("T", "F")),
    (78, ("R", "E"), ("R", "I")),
    (79, ("W", "A"), ("T", "E")),
    (80, ("W", "L"), ("T", "K")),
    (81, ("W", "D"), ("T", "B")),
    (82, ("W", "G"), ("T", "I")),
    (83, ("W", "H"), ("R", "J")),
    (84, ("W", "B"), ("T", "J")),
    (85, ("R", "K"), ("R", "L")),
    (86, ("W", "J"), ("R", "H")),
    (87, ("W", "K"), ("T", "L")),
    (88, ("R", "D"), ("R", "G")),
]
# Árvore a partir das oitavas: (num_jogo, vencedor_de_X, vencedor_de_Y)
R16 = [(89, 74, 77), (90, 73, 75), (91, 76, 78), (92, 79, 80),
       (93, 83, 84), (94, 81, 82), (95, 86, 88), (96, 85, 87)]
QF = [(97, 89, 90), (98, 91, 92), (99, 93, 94), (100, 95, 96)]
SF = [(101, 97, 98), (102, 99, 100)]
FINAL_NUM = 104       # vencedor 101 x vencedor 102
TERCEIRO_NUM = 103    # perdedor 101 x perdedor 102

NOME_FASE = {73: "Oitavas (R32)", 89: "16-avos (R16)", 97: "Quartas",
             101: "Semifinal", 103: "3º lugar", 104: "FINAL"}


def fase_do_jogo(num):
    if 73 <= num <= 88:
        return "Oitavas (R32)"
    if 89 <= num <= 96:
        return "16-avos (R16)"
    if 97 <= num <= 100:
        return "Quartas"
    if num in (101, 102):
        return "Semifinal"
    if num == 103:
        return "3º lugar"
    return "FINAL"


# ==========================================================
# 1) LEITURA DOS DADOS
# ==========================================================
def norm(nome):
    if nome is None:
        return None
    return NORMALIZA_NOME.get(str(nome).strip(), str(nome).strip())


def ler_painel(wb):
    ws = wb["Painel"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    jogos = []
    for r in linhas[1:]:
        if r[h["gols_pro"]] is None or r[h["gols_contra"]] is None:
            continue
        jogos.append({
            "data": str(r[h["data_jogo"]]),
            "time": norm(r[h["time"]]),
            "adv": norm(r[h["adversario"]]),
            "gp": int(r[h["gols_pro"]]),
            "gc": int(r[h["gols_contra"]]),
            "chutes_alvo": _num(r[h.get("fin_no_gol")]) if "fin_no_gol" in h else None,
        })
    return jogos


def ler_grupos(wb):
    ws = wb["Classificacao"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    grupo = {}
    for r in linhas[1:]:
        sel = norm(r[h["selecao"]])
        g = str(r[h["grupo"]]).replace("Group ", "").replace("Grupo ", "").strip()
        if sel:
            grupo[sel] = g
    return grupo


def ler_datas_mata_mata(wb):
    """Pega as datas previstas de cada fase na aba 'Partidas'."""
    ws = wb["Partidas"]
    linhas = list(ws.iter_rows(values_only=True))
    h = {nome: i for i, nome in enumerate(linhas[0])}
    datas = defaultdict(list)
    mapa = {"LAST_32": "Oitavas (R32)", "LAST_16": "16-avos (R16)",
            "QUARTER_FINALS": "Quartas", "SEMI_FINALS": "Semifinal",
            "THIRD_PLACE": "3º lugar", "FINAL": "FINAL"}
    for r in linhas[1:]:
        fase = mapa.get(str(r[h["fase"]]))
        if fase:
            datas[fase].append(str(r[h["data_jogo"]]))
    for k in datas:
        datas[k].sort()
    return datas


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("%", "").strip())
    except ValueError:
        return None


# ==========================================================
# 2) CLASSIFICAÇÃO DOS GRUPOS (a partir dos resultados do Painel)
# ==========================================================
def classificar(jogos, grupo):
    tab = defaultdict(lambda: dict(J=0, V=0, E=0, D=0, GP=0, GC=0, Pts=0))
    for j in jogos[: ]:
        # só conta jogos da fase de grupos
        if j["data"][:16] >= INICIO_MATA_MATA:
            continue
        s = tab[j["time"]]
        s["J"] += 1; s["GP"] += j["gp"]; s["GC"] += j["gc"]
        if j["gp"] > j["gc"]:
            s["V"] += 1; s["Pts"] += 3
        elif j["gp"] < j["gc"]:
            s["D"] += 1
        else:
            s["E"] += 1; s["Pts"] += 1

    def chave(t):
        s = tab[t]
        return (s["Pts"], s["GP"] - s["GC"], s["GP"])

    posicoes = {}     # (pos, grupo) -> time   ex: ('W','A') -> 'Mexico'
    terceiros = []    # lista de (time, grupo, stats)
    for g in sorted(set(grupo.values())):
        times = sorted([t for t in tab if grupo.get(t) == g],
                       key=chave, reverse=True)
        if len(times) >= 1:
            posicoes[("W", g)] = times[0]
        if len(times) >= 2:
            posicoes[("R", g)] = times[1]
        if len(times) >= 3:
            terceiros.append((times[2], g, tab[times[2]]))
    return tab, posicoes, terceiros


def melhores_terceiros(terceiros):
    """Ranqueia os 12 terceiros e devolve os 8 melhores (grupo -> time)."""
    ordenados = sorted(
        terceiros,
        key=lambda x: (x[2]["Pts"], x[2]["GP"] - x[2]["GC"], x[2]["GP"]),
        reverse=True,
    )
    melhores = ordenados[:8]
    return {g: t for (t, g, _) in melhores}


# ==========================================================
# 3) FORÇA DAS SELEÇÕES (Elo + ataque/defesa Poisson com xG)
# ==========================================================
def calcular_forcas(jogos):
    # ---- 3a) Elo, começando da FORÇA REAL pré-Copa e ajustando jogo a jogo ----
    times = set()
    for j in jogos:
        times.add(j["time"]); times.add(j["adv"])
    elo = {t: float(ELO_INICIAL.get(t, DEFAULT_ELO)) for t in times}
    K = 20                # baixo de propósito: 3 jogos não devem mudar muito o rating
    vistos = set()
    for j in sorted(jogos, key=lambda x: x["data"]):
        par = tuple(sorted([j["time"], j["adv"]]) + [j["data"]])
        if par in vistos:      # cada jogo aparece 2x no Painel (casa/fora)
            continue
        vistos.add(par)
        a, b = j["time"], j["adv"]
        ea = 1 / (1 + 10 ** ((elo[b] - elo[a]) / 400))
        if j["gp"] > j["gc"]:
            sa = 1.0
        elif j["gp"] < j["gc"]:
            sa = 0.0
        else:
            sa = 0.5
        mult = 1 + math.log(1 + abs(j["gp"] - j["gc"]))  # goleada pesa mais
        delta = K * mult * (sa - ea)
        elo[a] += delta
        elo[b] -= delta

    # ---- 3b) "gols ajustados" = mistura de gols reais com proxy de xG ----
    def gol_ajustado(g, chutes_alvo):
        if chutes_alvo is None:
            return float(g)
        xg = chutes_alvo * XG_POR_CHUTE_NO_ALVO
        return PESO_GOLS_REAIS * g + (1 - PESO_GOLS_REAIS) * xg

    media_liga = (sum(gol_ajustado(j["gp"], j["chutes_alvo"]) for j in jogos)
                  / max(len(jogos), 1))

    # prior de cada time vindo do Elo (Elo alto -> ataque maior, defesa menor)
    elo_medio = sum(elo.values()) / len(elo)
    prior = {}
    for t in elo:
        s = math.exp((elo[t] - elo_medio) / ESCALA_FORCA)   # força geral
        prior[t] = {"atk": math.sqrt(s), "dfs": 1 / math.sqrt(s)}

    # ataque/defesa por iteração, "puxando" para o prior do Elo (encolhimento)
    atk = {t: prior[t]["atk"] for t in elo}
    dfs = {t: prior[t]["dfs"] for t in elo}
    PESO_PRIOR = 3.0   # peso do prior (em "jogos equivalentes")

    for _ in range(60):
        novo_atk, novo_dfs = {}, {}
        for t in elo:
            jg = [j for j in jogos if j["time"] == t]
            n = len(jg)
            gp = sum(gol_ajustado(j["gp"], j["chutes_alvo"]) for j in jg)
            esperado_atk = sum(media_liga * dfs[j["adv"]] for j in jg)
            gc = sum(gol_ajustado(j["gc"], None) for j in jg)
            esperado_dfs = sum(media_liga * atk[j["adv"]] for j in jg)
            # encolhimento bayesiano simples para o prior do Elo
            novo_atk[t] = ((gp + PESO_PRIOR * prior[t]["atk"] * media_liga) /
                           (esperado_atk + PESO_PRIOR * media_liga))
            novo_dfs[t] = ((gc + PESO_PRIOR * prior[t]["dfs"] * media_liga) /
                           (esperado_dfs + PESO_PRIOR * media_liga))
        atk, dfs = novo_atk, novo_dfs

    return {"elo": dict(elo), "atk": atk, "dfs": dfs,
            "media_liga": media_liga, "prior": prior}


# ==========================================================
# 4) PREVISÃO DE UM JOGO (Poisson)
# ==========================================================
def _poisson(lmbda, k):
    return math.exp(-lmbda) * lmbda ** k / math.factorial(k)


def prever_jogo(a, b, F, max_gols=8):
    """Devolve probabilidades, placar mais provável e gols esperados."""
    la = F["media_liga"] * F["atk"].get(a, 1) * F["dfs"].get(b, 1)
    lb = F["media_liga"] * F["atk"].get(b, 1) * F["dfs"].get(a, 1)
    pa = pe = pb = 0.0
    melhor_p, placar = -1, (0, 0)
    for x in range(max_gols + 1):
        for y in range(max_gols + 1):
            p = _poisson(la, x) * _poisson(lb, y)
            if x > y:
                pa += p
            elif x == y:
                pe += p
            else:
                pb += p
            if p > melhor_p:
                melhor_p, placar = p, (x, y)
    # no mata-mata empate vai para prorrogação/pênaltis -> divide pela força
    forca_a = F["atk"].get(a, 1) / F["dfs"].get(a, 1)
    forca_b = F["atk"].get(b, 1) / F["dfs"].get(b, 1)
    p_pen_a = forca_a / (forca_a + forca_b)
    p_avanca_a = pa + pe * p_pen_a
    return {"la": la, "lb": lb, "pa": pa, "pe": pe, "pb": pb,
            "placar": placar, "p_avanca_a": p_avanca_a}


def resultado_real(a, b, jogos):
    """Se o jogo do mata-mata já foi jogado (está no Painel), devolve o vencedor."""
    for j in jogos:
        if j["data"][:16] < INICIO_MATA_MATA:
            continue
        if {j["time"], j["adv"]} == {a, b}:
            if j["gp"] == j["gc"]:
                continue          # empate no tempo normal: não dá pra saber o pênalti
            return j["time"] if j["gp"] > j["gc"] else j["adv"]
    return None


# ==========================================================
# 5) MONTE CARLO: chance de cada fase e de ser campeão
# ==========================================================
def simular_torneio(slots_r32, F, jogos, n=N_SIMULACOES):
    cache = {}

    def prob(a, b):
        key = (a, b)
        if key not in cache:
            cache[key] = prever_jogo(a, b, F)["p_avanca_a"]
        return cache[key]

    contagem = defaultdict(lambda: defaultdict(int))  # time -> fase -> vezes
    for _ in range(n):
        venc = {}
        # R32
        for (num, a, b) in slots_r32:
            r = resultado_real(a, b, jogos)
            if r is None:
                r = a if random.random() < prob(a, b) else b
            venc[num] = r
            contagem[a]["Oitavas (R32)"]  # garante chave
            contagem[b]["Oitavas (R32)"]
            contagem[r]["16-avos (R16)"] += 1
        # rounds seguintes
        for fase, jogos_fase, nome in [
            (R16, R16, "Quartas"), (QF, QF, "Semifinal"), (SF, SF, "FINAL")]:
            for (num, x, y) in jogos_fase:
                a, b = venc[x], venc[y]
                r = resultado_real(a, b, jogos)
                if r is None:
                    r = a if random.random() < prob(a, b) else b
                venc[num] = r
                if nome != "FINAL":
                    contagem[r][nome] += 1
                else:
                    contagem[a]["FINAL"]  # finalistas
                    contagem[b]["FINAL"]
        # final
        a, b = venc[101], venc[102]
        contagem[a]["FINAL"] += 0
        contagem[b]["FINAL"] += 0
        contagem[a].setdefault("Finalista", 0); contagem[a]["Finalista"] += 1
        contagem[b].setdefault("Finalista", 0); contagem[b]["Finalista"] += 1
        campeao = resultado_real(a, b, jogos) or (a if random.random() < prob(a, b) else b)
        contagem[campeao].setdefault("Campeão", 0)
        contagem[campeao]["Campeão"] += 1

    probs = {}
    for t, d in contagem.items():
        probs[t] = {
            "R16": d.get("16-avos (R16)", 0) / n,
            "QF": d.get("Quartas", 0) / n,
            "SF": d.get("Semifinal", 0) / n,
            "Final": d.get("Finalista", 0) / n,
            "Campeao": d.get("Campeão", 0) / n,
        }
    return probs


# ==========================================================
# 6) CHAVEAMENTO "FAVORITO AVANÇA" (para mostrar o caminho previsto)
# ==========================================================
def chave_favoritos(slots_r32, F, jogos):
    venc, jogos_prev = {}, []

    def joga(num, a, b):
        pr = prever_jogo(a, b, F)
        real = resultado_real(a, b, jogos)
        if real:
            r, status = real, "JOGADO"
        else:
            r = a if pr["p_avanca_a"] >= 0.5 else b
            status = "previsto"
        jogos_prev.append({"num": num, "fase": fase_do_jogo(num),
                           "a": a, "b": b, "pr": pr, "venc": r, "status": status})
        return r

    for (num, a, b) in slots_r32:
        venc[num] = joga(num, a, b)
    for grupo_fase in (R16, QF, SF):
        for (num, x, y) in grupo_fase:
            venc[num] = joga(num, venc[x], venc[y])
    # 3º lugar = perdedores das semis
    p101 = SF[0][1], SF[0][2]
    perd_sf = []
    for (num, x, y) in SF:
        a, b = venc[x], venc[y]
        perd_sf.append(a if venc[num] == b else b)
    venc[TERCEIRO_NUM] = joga(TERCEIRO_NUM, perd_sf[0], perd_sf[1])
    # final
    venc[FINAL_NUM] = joga(FINAL_NUM, venc[101], venc[102])
    return jogos_prev, venc


# ==========================================================
# 7) ESCREVER AS ABAS
# ==========================================================
AZUL = "1F4E78"; AZUL_CLARO = "DDEBF7"; VERDE = "C6EFCE"; CINZA = "F2F2F2"


def estilo_cabecalho(ws, ncols, linha=1):
    fill = PatternFill("solid", start_color=AZUL)
    fonte = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=linha, column=c)
        cell.fill = fill; cell.font = fonte
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bd
    ws.row_dimensions[linha].height = 22


def pct(x):
    return f"{100*x:.1f}%"


def escrever_predicoes(wb, jogos_prev, probs):
    if "Predicoes_MataMata" in wb.sheetnames:
        del wb["Predicoes_MataMata"]
    ws = wb.create_sheet("Predicoes_MataMata")

    arial = Font(name="Arial", size=11)
    bold = Font(name="Arial", size=11, bold=True)

    ws["A1"] = "CHAVEAMENTO PREVISTO — COPA 2026"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=AZUL)
    ws["A2"] = ("'JOGADO' = resultado real já no Painel | 'previsto' = predição do modelo "
                "(Poisson + Elo + proxy de xG). Empate no mata-mata é resolvido pela força "
                "do time (pênaltis).")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")

    cab = ["Jogo", "Fase", "Mandante", "Visitante", "Placar provável",
           "P(Mandante)", "Empate", "P(Visitante)", "Quem avança", "Status"]
    ws.append([])  # linha 3 vazia
    ws.append(cab)
    linha_cab = ws.max_row
    estilo_cabecalho(ws, len(cab), linha_cab)

    for jp in jogos_prev:
        pr = jp["pr"]
        ws.append([
            jp["num"], jp["fase"], jp["a"], jp["b"],
            f"{pr['placar'][0]}-{pr['placar'][1]}",
            pct(pr["pa"]), pct(pr["pe"]), pct(pr["pb"]),
            jp["venc"], jp["status"],
        ])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = arial
            ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r, 3).alignment = Alignment(horizontal="left")
        ws.cell(r, 4).alignment = Alignment(horizontal="left")
        ws.cell(r, 9).font = bold
        if jp["status"] == "JOGADO":
            for c in range(1, len(cab) + 1):
                ws.cell(r, c).fill = PatternFill("solid", start_color=CINZA)
        # destaca a FINAL
        if jp["num"] == FINAL_NUM:
            for c in range(1, len(cab) + 1):
                ws.cell(r, c).fill = PatternFill("solid", start_color=VERDE)
                ws.cell(r, c).font = bold

    larguras = [6, 16, 20, 20, 14, 12, 9, 12, 20, 10]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{linha_cab + 1}"

    # ---- tabela de probabilidades por seleção ----
    col0 = len(cab) + 2   # começa 1 coluna depois
    ws.cell(linha_cab - 1, col0, "PROBABILIDADES (Monte Carlo)").font = bold
    cab2 = ["Seleção", "Oitavas→R16", "→Quartas", "→Semi", "→Final", "Campeão"]
    for i, t in enumerate(cab2):
        ws.cell(linha_cab, col0 + i, t)
    estilo_cabecalho_intervalo(ws, linha_cab, col0, len(cab2))

    ranking = sorted(probs.items(), key=lambda kv: kv[1]["Campeao"], reverse=True)
    r = linha_cab + 1
    for time, p in ranking:
        vals = [time, pct(p["R16"]), pct(p["QF"]), pct(p["SF"]),
                pct(p["Final"]), pct(p["Campeao"])]
        for i, v in enumerate(vals):
            cell = ws.cell(r, col0 + i, v)
            cell.font = arial
            cell.alignment = Alignment(horizontal="center" if i else "left")
        # barra de cor na coluna campeão
        intensidade = min(p["Campeao"] * 3, 1)
        if intensidade > 0.02:
            ws.cell(r, col0 + 5).fill = PatternFill(
                "solid", start_color=cor_gradiente(intensidade))
        r += 1
    larg2 = [20, 12, 11, 10, 10, 11]
    for i, w in enumerate(larg2):
        ws.column_dimensions[get_column_letter(col0 + i)].width = w


def estilo_cabecalho_intervalo(ws, linha, col0, ncols):
    fill = PatternFill("solid", start_color=AZUL)
    fonte = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for c in range(col0, col0 + ncols):
        cell = ws.cell(linha, c)
        cell.fill = fill; cell.font = fonte
        cell.alignment = Alignment(horizontal="center", vertical="center")


def cor_gradiente(x):
    # de azul claro para azul forte
    r = int(0xDD + (0x1F - 0xDD) * x)
    g = int(0xEB + (0x4E - 0xEB) * x)
    b = int(0xF7 + (0x78 - 0xF7) * x)
    return f"{r:02X}{g:02X}{b:02X}"


def escrever_proximos(wb, jogos_prev, datas_fase):
    if "Proximos_Jogos" in wb.sheetnames:
        del wb["Proximos_Jogos"]
    ws = wb.create_sheet("Proximos_Jogos")

    arial = Font(name="Arial", size=11)
    bold = Font(name="Arial", size=11, bold=True)
    ws["A1"] = "PRÓXIMOS JOGOS + PREDIÇÃO"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=AZUL)
    ws["A2"] = "Jogos do mata-mata que ainda não foram jogados, em ordem de fase."
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="808080")

    cab = ["Data prevista", "Fase", "Jogo", "Placar provável",
           "P(time 1)", "Empate", "P(time 2)", "Favorito", "Confiança"]
    ws.append([]); ws.append(cab)
    linha_cab = ws.max_row
    estilo_cabecalho(ws, len(cab), linha_cab)

    # data por fase (pega a 1ª data daquela fase como referência)
    contador_fase = defaultdict(int)
    proximos = [jp for jp in jogos_prev if jp["status"] != "JOGADO"]
    for jp in proximos:
        pr = jp["pr"]
        fase = jp["fase"]
        lista = datas_fase.get(fase, [])
        idx = contador_fase[fase]
        data = lista[idx][:16] if idx < len(lista) else "a definir"
        contador_fase[fase] += 1
        conf = max(pr["pa"], pr["pb"])
        ws.append([
            data, fase, f"{jp['a']} x {jp['b']}",
            f"{pr['placar'][0]}-{pr['placar'][1]}",
            pct(pr["pa"]), pct(pr["pe"]), pct(pr["pb"]),
            jp["venc"], pct(conf),
        ])
        r = ws.max_row
        for c in range(1, len(cab) + 1):
            ws.cell(r, c).font = arial
            ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r, 3).alignment = Alignment(horizontal="left")
        ws.cell(r, 8).font = bold
        if conf >= 0.6:
            ws.cell(r, 9).fill = PatternFill("solid", start_color=VERDE)

    larguras = [16, 16, 30, 14, 11, 9, 11, 20, 11]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{linha_cab + 1}"


# ==========================================================
# MAIN
# ==========================================================
def main():
    try:
        wb = load_workbook(PLANILHA)
    except FileNotFoundError:
        print(f"Não achei '{PLANILHA}'. Deixe na mesma pasta do script.")
        return

    jogos = ler_painel(wb)
    grupo = ler_grupos(wb)
    datas_fase = ler_datas_mata_mata(wb)
    print(f"Jogos lidos do Painel: {len(jogos)//2}")

    tab, posicoes, terceiros = classificar(jogos, grupo)
    melhores = melhores_terceiros(terceiros)
    print("8 melhores terceiros (grupos):", sorted(melhores.keys()))

    # preenche os 32 times nas vagas do R32
    def resolve(vaga):
        tipo, g = vaga
        if tipo in ("W", "R"):
            return posicoes.get((tipo, g))
        return melhores.get(g)  # terceiro

    slots_r32 = []
    for (num, va, vb) in R32:
        a, b = resolve(va), resolve(vb)
        if a is None or b is None:
            print(f"  ! vaga vazia no jogo {num}: {va}={a} {vb}={b}")
        slots_r32.append((num, a, b))

    F = calcular_forcas(jogos)

    print("\nTop 8 por Elo:")
    for t, e in sorted(F["elo"].items(), key=lambda kv: kv[1], reverse=True)[:8]:
        print(f"  {t:<22} Elo {e:6.0f}")

    jogos_prev, _ = chave_favoritos(slots_r32, F, jogos)
    probs = simular_torneio(slots_r32, F, jogos)

    escrever_predicoes(wb, jogos_prev, probs)
    escrever_proximos(wb, jogos_prev, datas_fase)
    wb.save(PLANILHA)

    print("\n=== CAMINHO PREVISTO ATÉ A FINAL ===")
    for jp in jogos_prev:
        if jp["num"] in (101, 102, 103, 104):
            print(f"  {jp['fase']:<12} {jp['a']} x {jp['b']}  -> {jp['venc']}")
    campeao = max(probs.items(), key=lambda kv: kv[1]["Campeao"])
    print(f"\nFavorito ao título: {campeao[0]} ({pct(campeao[1]['Campeao'])})")
    print(f"\nAbas criadas: 'Predicoes_MataMata' e 'Proximos_Jogos'. Planilha salva.")


if __name__ == "__main__":
    main()
